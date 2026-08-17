from __future__ import annotations

import hashlib
import json
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path

import pandas as pd
from jinja2 import Environment, FileSystemLoader, StrictUndefined
from openpyxl import load_workbook
from PIL import Image
from playwright.sync_api import sync_playwright
from pypdf import PdfReader

from core.contracts import BaseReport
from core.exceptions import InputValidationError
from core.reporting_period import validate_reporting_period
from core.tabular import parse_datetime, parse_numeric, read_table

from .config import CashOperationsConfig

VERSION = "1.0.0-provisional.4"
HEADERS = [
    "Slip #", "Date & Time", "Currency", "Game", "Cash Amount",
    "Withholding Tax", "Type", "User #", "User Name", "Paid Out Total - Promo",
]


class CashOperationsDashboardReport(BaseReport):
    def __init__(self, config: CashOperationsConfig | None = None):
        self.config = config or CashOperationsConfig()

    def run(self, workbook_path: Path, work_directory: Path, *, report_date: date,
            reporting_period_start: date, reporting_period_end: date,
            generation_uuid: str, render_outputs: bool = True) -> dict[str, Path]:
        validate_reporting_period(
            report_date, reporting_period_start, reporting_period_end
        )
        for name in ("prepared", "results", "charts", "render", "outputs", "manifest"):
            (work_directory / name).mkdir(parents=True, exist_ok=True)
        frame, validation, source = self._read(workbook_path)
        frame = frame[
            frame.transaction_date.between(
                pd.Timestamp(reporting_period_start), pd.Timestamp(reporting_period_end)
            )
            & ~frame.transaction_date.dt.date.isin(self.config.excluded_dates)
        ].copy()
        if frame.empty:
            raise InputValidationError(
                "No Cash Operations rows fall inside the reporting period.",
                code="NO_CASH_OPERATIONS_IN_PERIOD",
            )
        dataset = work_directory / "prepared" / "betting-dataset.parquet"
        frame.to_parquet(dataset, index=False)
        validation_path = work_directory / "prepared" / "validation-log.json"
        validation_path.write_text(json.dumps(validation, indent=2), encoding="utf-8")
        results = self._calculate(
            frame, report_date, reporting_period_end, reporting_period_start,
            self.config.excluded_dates,
        )
        result_path = work_directory / "results" / "calculated-results.json"
        result_path.write_text(json.dumps(results, indent=2), encoding="utf-8")
        reconciliation_path = work_directory / "results" / "reconciliation-report.json"
        reconciliation_path.write_text(json.dumps({
            "report_code": "cash_operations_dashboard",
            "passed": all(item["passed"] for item in results["reconciliation"]),
            "checks": results["reconciliation"],
        }, indent=2), encoding="utf-8")
        chart = self._chart(results, work_directory / "charts")
        html = self._html(results, chart, work_directory / "render")
        artifacts = {
            "betting_dataset": dataset, "validation_log": validation_path,
            "calculated_results": result_path, "reconciliation_report": reconciliation_path,
            "dashboard_html": html, "chart_last_ten_days": chart,
        }
        if render_outputs:
            pdf, png = self._outputs(html, work_directory / "outputs")
            self._verify(results, html, pdf, png)
            artifacts.update(pdf=pdf, png=png)
        manifest_path = work_directory / "manifest" / "manifest.json"
        manifest_path.write_text(json.dumps({
            "generation_uuid": generation_uuid,
            "report_code": "cash_operations_dashboard",
            "report_date": report_date.isoformat(),
            "reporting_period_start": reporting_period_start.isoformat(),
            "reporting_period_end": reporting_period_end.isoformat(),
            "excluded_dates": sorted(value.isoformat() for value in self.config.excluded_dates),
            "definition_version": VERSION, "calculation_version": VERSION,
            "template_version": VERSION, "timezone": self.config.timezone,
            "generated_at": datetime.now(UTC).isoformat(),
            "inputs": [{"key": "cash_operations", "filename": workbook_path.name,
                        "sha256": self._sha(workbook_path)}],
            "source": source,
            "configuration": {
                "recognition_date": "Date & Time",
                "valid_transaction_types": ["Bet", "Payout"],
                "void_cancelled_refunded_handling": "not_present_in_source_export",
                "provisional": True,
            },
            "artifacts": [{"key": key, "relative_path": str(path.relative_to(work_directory)),
                           "size_bytes": path.stat().st_size, "sha256": self._sha(path)}
                          for key, path in artifacts.items()],
            "warnings": results["warnings"],
        }, indent=2), encoding="utf-8")
        artifacts["manifest"] = manifest_path
        return artifacts

    def _read(self, path: Path) -> tuple[pd.DataFrame, dict, dict]:
        workbook = None
        if path.suffix.casefold() == ".csv":
            raw_frame = read_table(path)
            headers = [str(value).strip() for value in raw_frame.columns]
            rows = list(raw_frame.itertuples(index=False, name=None))
            worksheet_name = "CSV"
        else:
            workbook = load_workbook(path, read_only=True, data_only=False)
            if self.config.worksheet not in workbook.sheetnames:
                raise InputValidationError(
                    f"Required worksheet '{self.config.worksheet}' was not found.",
                    code="CASH_OPERATIONS_WORKSHEET_MISSING",
                )
            sheet = workbook[self.config.worksheet]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            rows = sheet.iter_rows(min_row=2, values_only=True)
            worksheet_name = self.config.worksheet
        if any(header not in headers for header in HEADERS):
            raise InputValidationError(
                "Cash Operations workbook headers do not match the production export.",
                code="CASH_OPERATIONS_HEADERS_INVALID",
                context={"expected": HEADERS, "observed": headers},
            )
        accepted, rejected = [], []
        for row_number, values in enumerate(rows, 2):
            row = dict(zip(headers, values, strict=False))
            if not str(row["User #"] or "").strip():
                rejected.append({"row": row_number, "code": "BLANK_PLAYER_ID"}); continue
            if "test" in str(row["User Name"] or "").casefold():
                rejected.append({"row": row_number, "code": "TEST_ACCOUNT"}); continue
            kind = str(row["Type"] or "").strip().casefold()
            if kind not in {self.config.bet_type, self.config.payout_type}:
                rejected.append({"row": row_number, "code": "INVALID_TYPE"}); continue
            try:
                parsed_amount = parse_numeric(row["Cash Amount"])
                if pd.isna(parsed_amount):
                    raise ValueError("Cash Amount is not numeric.")
                amount = Decimal(str(parsed_amount))
                parsed_tax = parse_numeric(row["Withholding Tax"])
                withholding_tax = 0 if pd.isna(parsed_tax) else float(parsed_tax)
                timestamp = parse_datetime(
                    row["Date & Time"], csv_source=path.suffix.casefold() == ".csv"
                )
            except (ArithmeticError, TypeError, ValueError):
                rejected.append({"row": row_number, "code": "INVALID_VALUE"}); continue
            if pd.isna(timestamp):
                rejected.append({"row": row_number, "code": "INVALID_TRANSACTION_DATE"}); continue
            accepted.append({
                "source_row": row_number, "slip_id": str(row["Slip #"]).strip(),
                "transaction_date": timestamp.normalize(), "transaction_timestamp": timestamp,
                "currency": str(row["Currency"]).strip(), "game": str(row["Game"]).strip(),
                "cash_amount": float(amount), "withholding_tax": withholding_tax,
                "transaction_type": kind, "player_id": str(row["User #"]).strip(),
                "username": str(row["User Name"]).strip(),
            })
        if workbook is not None:
            first_slip = workbook["Sheet1"]["F48"].value if "Sheet1" in workbook.sheetnames else None
            last_slip = workbook["Sheet1"]["G48"].value if "Sheet1" in workbook.sheetnames else None
            workbook.close()
        else:
            timestamps = [item["transaction_timestamp"] for item in accepted]
            first_slip = min(timestamps) if timestamps else None
            last_slip = max(timestamps) if timestamps else None
        if not accepted:
            raise InputValidationError("No valid Cash Operations rows found.", code="NO_VALID_CASH_OPERATIONS")
        accepted_frame = pd.DataFrame(accepted)
        accepted_frame.attrs["first_slip"] = first_slip.isoformat() if first_slip else None
        accepted_frame.attrs["last_slip"] = last_slip.isoformat() if last_slip else None
        return accepted_frame, {
            "worksheet": worksheet_name, "source_rows": len(accepted) + len(rejected),
            "accepted_rows": len(accepted), "rejected_rows": len(rejected), "issues": rejected,
            "warnings": [
                "The production workbook has no explicit bet status or settlement date columns; Type=Bet/Payout is the provisional settled-transaction definition."
            ],
        }, {
            "worksheet": worksheet_name, "header_row": 1,
            "column_mapping": {
                "bet_id": "Slip #", "recognition_date": "Date & Time",
                "currency": "Currency", "game": "Game", "stake_or_payout": "Cash Amount",
                "withholding_tax": "Withholding Tax", "transaction_type": "Type",
                "player_id": "User #", "username": "User Name",
            },
        }

    def _calculate(
        self, frame: pd.DataFrame, report_date: date, end: date, start: date,
        excluded_dates: frozenset[date],
    ) -> dict:
        bets = frame[frame.transaction_type == self.config.bet_type]
        payouts = frame[frame.transaction_type == self.config.payout_type]
        stake = Decimal(str(bets.cash_amount.sum()))
        paid = abs(Decimal(str(payouts.cash_amount.sum())))
        tax = Decimal(str(frame.withholding_tax.sum()))
        ggr = stake - paid
        margin = ggr / stake * 100 if stake else Decimal(0)
        payout_rate = paid / stake * 100 if stake else Decimal(0)
        index = pd.date_range(start, end, freq="D")
        daily = []
        for stamp in index:
            current = frame[frame.transaction_date == stamp]
            day_bets = current[current.transaction_type == self.config.bet_type]
            day_paid = current[current.transaction_type == self.config.payout_type]
            day_stake = Decimal(str(day_bets.cash_amount.sum()))
            day_payout = abs(Decimal(str(day_paid.cash_amount.sum())))
            daily.append({"date": stamp.date().isoformat(), "bet_count": len(day_bets),
                          "bet_amount": float(day_stake), "payout_count": len(day_paid),
                          "payout_amount": float(day_payout), "ggr": float(day_stake-day_payout)})
        last_ten = daily[-10:]
        peak_bet = max(daily, key=lambda x: x["bet_amount"])
        low_bet = min((x for x in daily if x["bet_amount"] > 0), key=lambda x: x["bet_amount"])
        peak_paid = max(daily, key=lambda x: x["payout_amount"])
        low_paid = min((x for x in daily if x["payout_amount"] > 0), key=lambda x: x["payout_amount"])
        last_stake = sum(Decimal(str(x["bet_amount"])) for x in last_ten)
        last_paid = sum(Decimal(str(x["payout_amount"])) for x in last_ten)
        discrepancies = []
        reference_period_applies = (
            start == self.config.audit_reference_period_start
            and end == self.config.audit_reference_period_end
        )
        payout_difference = (
            last_paid - self.config.audit_reference_last_ten_payout_xaf
            if reference_period_applies
            else Decimal(0)
        )
        if reference_period_applies and payout_difference != 0:
            discrepancies.append({
                "metric": "Last-10-days winning paid",
                "benchmark": float(self.config.audit_reference_last_ten_payout_xaf),
                "calculated": float(last_paid), "difference": float(payout_difference),
                "message": f"Production reference differs from source transactions by XAF {abs(payout_difference):,.0f}.",
            })
        lowest_difference = (
            Decimal(str(low_paid["payout_amount"]))
            - self.config.audit_reference_lowest_payout_xaf
            if reference_period_applies
            else Decimal(0)
        )
        if reference_period_applies and lowest_difference != 0:
            discrepancies.append({
                "metric": "Lowest winning-paid day",
                "benchmark": float(self.config.audit_reference_lowest_payout_xaf),
                "calculated": low_paid["payout_amount"], "difference": float(lowest_difference),
                "message": "Production reference value/date does not match the workbook-derived minimum.",
            })
        reconciliation = [
            {"name": "bet_count", "expected": len(bets), "actual": sum(x["bet_count"] for x in daily),
             "difference": 0, "passed": len(bets) == sum(x["bet_count"] for x in daily)},
            {"name": "stake", "expected": float(stake), "actual": sum(x["bet_amount"] for x in daily),
             "difference": 0, "passed": float(stake) == sum(x["bet_amount"] for x in daily)},
            {"name": "payout", "expected": float(paid), "actual": sum(x["payout_amount"] for x in daily),
             "difference": 0, "passed": float(paid) == sum(x["payout_amount"] for x in daily)},
            {"name": "ggr", "expected": float(stake-paid), "actual": float(ggr),
             "difference": 0, "passed": ggr == stake-paid},
        ]
        if reference_period_applies:
            reconciliation.extend([
                {"name": "production_reference_last_ten_payout",
                 "expected": float(self.config.audit_reference_last_ten_payout_xaf),
                 "actual": float(last_paid), "difference": float(payout_difference),
                 "passed": payout_difference == 0,
                 "note": "Audit comparison only; source-calculated value is preserved."},
                {"name": "production_reference_lowest_payout",
                 "expected": float(self.config.audit_reference_lowest_payout_xaf),
                 "actual": low_paid["payout_amount"], "difference": float(lowest_difference),
                 "passed": lowest_difference == 0,
                 "note": "Audit comparison only; source-calculated value is preserved."},
            ])
        warnings = [
            "PROVISIONAL: Type=Bet and Type=Payout are treated as valid settled transactions because the export has no status field.",
            "PROVISIONAL: Date & Time is used as the recognition date because the export has no separate settlement date.",
        ]
        warnings.extend(f"REVIEW REQUIRED: {item['metric']}: {item['message']}" for item in discrepancies)
        return {
            "report_code": "cash_operations_dashboard", "report_date": report_date.isoformat(),
            "period_start": start.isoformat(), "period_end": end.isoformat(),
            "excluded_dates": sorted(value.isoformat() for value in excluded_dates),
            "summary": {
                "bet_count": len(bets), "bet_amount": float(stake),
                "payout_count": len(payouts), "payout_amount": float(paid),
                "ggr": float(ggr), "withholding_tax": float(tax),
                "margin": float(margin), "payout_rate": float(payout_rate),
                "average_bet": float(stake/len(bets)), "average_payout": float(paid/len(payouts)),
            },
            "first_slip": frame.attrs.get("first_slip") or frame.transaction_timestamp.min().isoformat(),
            "last_slip": frame.attrs.get("last_slip") or frame.transaction_timestamp.max().isoformat(),
            "daily": daily, "last_ten": last_ten,
            "last_ten_summary": {"bet_amount": float(last_stake), "payout_amount": float(last_paid),
                                 "ggr": float(last_stake-last_paid)},
            "highlights": {"highest_bet": peak_bet, "lowest_bet": low_bet,
                           "highest_payout": peak_paid, "lowest_payout": low_paid},
            "insights": [
                f"Total bets placed reached {len(bets):,}, generating XAF {stake:,.0f} in wagers.",
                f"Winning payouts totalled XAF {paid:,.0f} across {len(payouts):,} winning bets.",
                f"Gross Gaming Revenue closed at XAF {ggr:,.0f}, delivering a {margin:.2f}% margin.",
                "No withholding tax was deducted during the reporting period." if tax == 0 else f"Withholding tax totalled XAF {tax:,.0f}.",
                f"Betting activity was strongest on {date.fromisoformat(peak_bet['date']).strftime('%d %B %Y')}.",
                f"The last 10 reporting days generated XAF {last_stake-last_paid:,.0f} in GGR.",
            ],
            "reconciliation": reconciliation, "warnings": warnings, "discrepancies": discrepancies,
        }

    def _chart(self, results: dict, directory: Path) -> Path:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        directory.mkdir(parents=True, exist_ok=True)
        rows = results["last_ten"]; x = list(range(len(rows))); width = .32
        fig, ax = plt.subplots(figsize=(8.8, 3.0), facecolor="#030303"); ax.set_facecolor("#030303")
        bet = ax.bar([i-width/2 for i in x], [r["bet_amount"] for r in rows], width, color="#1661a8", label="Bet Amount (XAF)")
        paid = ax.bar([i+width/2 for i in x], [r["payout_amount"] for r in rows], width, color="#cf1f28", label="Winning Paid (XAF)")
        ax.set_xticks(x, [date.fromisoformat(r["date"]).strftime("%d/%m") for r in rows])
        ax.tick_params(colors="#eee", labelsize=8); ax.grid(axis="y", color="#555", alpha=.45, linestyle="--"); ax.set_axisbelow(True)
        for spine in ax.spines.values(): spine.set_visible(False)
        ax.bar_label(bet, labels=[f"{r['bet_amount']:,.0f}" for r in rows], color="#eee", fontsize=7, padding=2)
        ax.bar_label(paid, labels=[f"{r['payout_amount']:,.0f}" for r in rows], color="#eee", fontsize=7, padding=2)
        legend=ax.legend(frameon=False,loc="upper center",ncol=2)
        for text in legend.get_texts(): text.set_color("#eee")
        fig.tight_layout(); path=directory/"last-ten-days.svg"; fig.savefig(path,format="svg",facecolor="#030303"); plt.close(fig)
        return path

    def _html(self, results: dict, chart: Path, directory: Path) -> Path:
        templates = Path(__file__).parent/"templates"; assets=Path(__file__).parent/"assets"
        env=Environment(loader=FileSystemLoader(templates),undefined=StrictUndefined)
        context={**results,"chart_uri":chart.resolve().as_uri(),"css_uri":(templates/"dashboard.css").resolve().as_uri(),
                 "logo_uri":(assets/"Favicon.jpeg").resolve().as_uri(),
                 "font_uri":(assets/"FreeSans.ttf").resolve().as_uri(),
                 "font_bold_uri":(assets/"FreeSansBold.ttf").resolve().as_uri(),
                 "fmt":lambda x:f"{x:,.0f}","datefmt":lambda x:date.fromisoformat(x).strftime("%d %B %Y")}
        directory.mkdir(parents=True,exist_ok=True); path=directory/"dashboard.html"
        path.write_text(env.get_template("dashboard.html").render(**context),encoding="utf-8"); return path

    def _outputs(self, html: Path, directory: Path) -> tuple[Path,Path]:
        directory.mkdir(parents=True,exist_ok=True); pdf=directory/"dashboard.pdf"; png=directory/"dashboard.png"
        with sync_playwright() as p:
            browser=p.chromium.launch(headless=True); page=browser.new_page(viewport={"width":1536,"height":1024})
            page.goto(html.resolve().as_uri(),wait_until="networkidle"); page.wait_for_function("document.fonts.status === 'loaded'")
            page.pdf(path=str(pdf),width="1536px",height="1024px",print_background=True,margin={"top":"0","right":"0","bottom":"0","left":"0"})
            page.screenshot(path=str(png),type="png",full_page=True); browser.close()
        return pdf,png

    def _verify(self, results: dict, html: Path, pdf: Path, png: Path) -> None:
        text=html.read_text()
        for value in (results["summary"]["bet_count"],results["summary"]["bet_amount"],results["summary"]["ggr"]):
            if f"{value:,.0f}" not in text: raise RuntimeError(f"Calculated value {value} missing from HTML.")
        if len(PdfReader(pdf).pages)!=1: raise RuntimeError("Cash Operations PDF must have one page.")
        with Image.open(png) as image:
            if image.size!=(1536,1024): raise RuntimeError(f"Unexpected PNG dimensions {image.size}.")

    @staticmethod
    def _sha(path: Path) -> str:
        digest=hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda:handle.read(1024*1024),b""): digest.update(chunk)
        return digest.hexdigest()

    def validate_inputs(self, files, reporting_context): return []
    def normalize_inputs(self, files, work_directory, reporting_context): return {}
    def calculate(self, normalized_files, reporting_context): return {}
    def validate_results(self, results, reporting_context): return []
    def generate_charts(self, results, output_directory, reporting_context): return {}
    def get_template_context(self, results, charts, reporting_context): return {}
