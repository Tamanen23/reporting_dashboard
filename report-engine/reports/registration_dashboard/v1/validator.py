import re
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Literal, cast

import pandas as pd
from openpyxl import load_workbook

from core.exceptions import InputValidationError
from core.tabular import parse_datetime, read_table

from .config import COLUMN_ALIASES, REQUIRED_FIELDS, RegistrationConfig
from .schemas import ValidationRecord


def _normalise_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _normalise_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _normalise_flag(value: Any) -> str:
    return _normalise_text(value).casefold()


@dataclass(frozen=True)
class ValidatedWorkbook:
    worksheet: str
    source_filename: str
    frame: pd.DataFrame
    issues: list[ValidationRecord]
    source_columns: dict[str, str]


class RegistrationWorkbookValidator:
    def __init__(self, config: RegistrationConfig):
        self.config = config

    def validate(
        self, workbook_path: Path, period_start: date, period_end: date, report_date: date
    ) -> ValidatedWorkbook:
        if not workbook_path.is_file():
            raise InputValidationError(
                "The User List workbook does not exist.",
                code="INPUT_FILE_NOT_FOUND",
                context={"path": str(workbook_path)},
            )
        if workbook_path.suffix.casefold() not in {".xlsx", ".csv"}:
            raise InputValidationError(
                "The User List input must be an XLSX workbook or CSV file.",
                code="INVALID_INPUT_EXTENSION",
            )
        if workbook_path.stat().st_size == 0:
            raise InputValidationError("The User List workbook is empty.", code="EMPTY_INPUT_FILE")

        workbook = None
        formula_rows: set[int] = set()
        records: list[dict[str, Any]] = []
        if workbook_path.suffix.casefold() == ".csv":
            worksheet_name = "CSV"
            source_frame = read_table(workbook_path)
            raw_headers = [_normalise_text(value) for value in source_frame.columns]
            for source_row_number, values in enumerate(source_frame.to_dict("records"), start=2):
                row = {header: values.get(header) for header in raw_headers}
                if any(value not in (None, "") for value in row.values()):
                    row["_source_row_number"] = source_row_number
                    records.append(row)
        else:
            workbook = load_workbook(workbook_path, read_only=True, data_only=False)
            worksheet_name = self.config.worksheet or workbook.sheetnames[0]
            if worksheet_name not in workbook.sheetnames:
                raise InputValidationError(
                    f"Worksheet '{worksheet_name}' was not found.",
                    code="WORKSHEET_NOT_FOUND",
                    context={"available_worksheets": workbook.sheetnames},
                )
            sheet = workbook[worksheet_name]
            rows = sheet.iter_rows(values_only=False)
            try:
                header_cells = next(rows)
            except StopIteration as exc:
                raise InputValidationError(
                    "The workbook contains no rows.", code="EMPTY_WORKSHEET"
                ) from exc
            raw_headers = [_normalise_text(cell.value) for cell in header_cells]

        normalised_headers = [_normalise_header(header) for header in raw_headers]
        duplicates = sorted(
            {
                header
                for header in normalised_headers
                if header and normalised_headers.count(header) > 1
            }
        )
        if duplicates:
            raise InputValidationError(
                "The worksheet contains duplicate column headings.",
                code="DUPLICATE_COLUMNS",
                context={"columns": duplicates},
            )

        source_columns = self._resolve_columns(raw_headers)
        missing = sorted(REQUIRED_FIELDS - source_columns.keys())
        if missing:
            raise InputValidationError(
                "The User List workbook is missing required columns.",
                code="MISSING_REQUIRED_COLUMNS",
                context={"missing_columns": missing, "observed_columns": raw_headers},
            )

        if workbook is not None:
            for source_row_number, cells in enumerate(rows, start=2):
                row = {
                    raw_headers[index]: cell.value
                    for index, cell in enumerate(cells)
                    if index < len(raw_headers)
                }
                if not any(value not in (None, "") for value in row.values()):
                    continue
                for source in source_columns.values():
                    index = raw_headers.index(source)
                    if index < len(cells) and cells[index].data_type == "f":
                        formula_rows.add(source_row_number)
                        row[source] = None
                row["_source_row_number"] = source_row_number
                records.append(row)
            workbook.close()

        if not records:
            raise InputValidationError("The raw worksheet has no data rows.", code="EMPTY_DATASET")

        canonical_rows = []
        issues: list[ValidationRecord] = []
        now = datetime.now(UTC)
        for row in records:
            canonical_row = {field: row.get(source) for field, source in source_columns.items()}
            canonical_row["_source_row_number"] = row["_source_row_number"]
            identifier = _normalise_text(canonical_row.get("player_id")) or None
            if row["_source_row_number"] in formula_rows:
                issues.append(
                    self._issue(
                        workbook_path,
                        worksheet_name,
                        row,
                        identifier,
                        "FORMULA_VALUE_REJECTED",
                        "Formula cells are not accepted in mapped source fields.",
                        now,
                    )
                )
                continue
            player_id = _normalise_text(canonical_row.get("player_id"))
            if not player_id or player_id.casefold() in {"nan", "none", "null"}:
                issues.append(
                    self._issue(
                        workbook_path,
                        worksheet_name,
                        row,
                        None,
                        "BLANK_PLAYER_ID",
                        "Player ID is blank or invalid.",
                        now,
                    )
                )
                continue
            username = _normalise_text(canonical_row.get("username"))
            if "test" in username.casefold():
                issues.append(
                    self._issue(
                        workbook_path,
                        worksheet_name,
                        row,
                        player_id,
                        "TEST_ACCOUNT",
                        "Username contains 'test' (case-insensitive).",
                        now,
                    )
                )
                continue
            parsed_date = parse_datetime(
                cast(Any, canonical_row.get("registration_date")),
                csv_source=workbook_path.suffix.casefold() == ".csv",
            )
            if pd.isna(parsed_date):
                issues.append(
                    self._issue(
                        workbook_path,
                        worksheet_name,
                        row,
                        player_id,
                        "INVALID_REGISTRATION_DATE",
                        "Registration date cannot be parsed.",
                        now,
                    )
                )
                continue
            registration_date = parsed_date.date()
            if registration_date < period_start or registration_date > period_end:
                issues.append(
                    self._issue(
                        workbook_path,
                        worksheet_name,
                        row,
                        player_id,
                        "OUTSIDE_REPORTING_PERIOD",
                        "Registration date is outside the selected reporting period.",
                        now,
                    )
                )
                continue
            if registration_date in self.config.excluded_dates or (
                registration_date == report_date and not self.config.include_report_date
            ):
                issues.append(
                    self._issue(
                        workbook_path,
                        worksheet_name,
                        row,
                        player_id,
                        "EXCLUDED_REPORTING_DATE",
                        "Registration date is excluded by reporting configuration.",
                        now,
                    )
                )
                continue

            canonical_row.update(
                {
                    "player_id": player_id,
                    "username": username,
                    "registration_date": registration_date,
                    "registration_completed": _normalise_flag(
                        canonical_row.get("registration_completed")
                    )
                    in self.config.completed_values,
                    "account_status": _normalise_flag(canonical_row.get("account_status")),
                    "identity_status": _normalise_flag(canonical_row.get("identity_status")),
                    "location_status": _normalise_flag(canonical_row.get("location_status")),
                    "is_disabled": _normalise_flag(canonical_row.get("disabled_status"))
                    in self.config.disabled_values,
                    "is_deleted": _normalise_flag(canonical_row.get("deleted_status"))
                    in self.config.deleted_values,
                    "last_deposit_date": parse_datetime(
                        cast(Any, canonical_row.get("last_deposit_date")),
                        csv_source=workbook_path.suffix.casefold() == ".csv",
                    ),
                    "email": _normalise_text(canonical_row.get("email")),
                    "mobile_number": _normalise_text(canonical_row.get("mobile_number")),
                    "country": _normalise_text(canonical_row.get("country")),
                    "currency": _normalise_text(canonical_row.get("currency")),
                    "auth_type": _normalise_text(canonical_row.get("auth_type")),
                    "date_of_birth": parse_datetime(
                        cast(Any, canonical_row.get("date_of_birth")),
                        csv_source=workbook_path.suffix.casefold() == ".csv",
                    ),
                    "tags": _normalise_text(canonical_row.get("tags")),
                    "extra_data": _normalise_text(canonical_row.get("extra_data")),
                    "last_login": parse_datetime(
                        cast(Any, canonical_row.get("last_login")),
                        csv_source=workbook_path.suffix.casefold() == ".csv",
                    ),
                    "balance": canonical_row.get("balance"),
                    "promo_balance": canonical_row.get("promo_balance"),
                    "promo_code": _normalise_text(canonical_row.get("promo_code")),
                    "pending_transactions": canonical_row.get("pending_transactions"),
                }
            )
            if canonical_row["is_deleted"] and self.config.exclude_deleted_accounts:
                issues.append(
                    self._issue(
                        workbook_path,
                        worksheet_name,
                        row,
                        player_id,
                        "DELETED_ACCOUNT",
                        "Deleted account excluded by provisional configuration.",
                        now,
                    )
                )
                continue
            canonical_rows.append(canonical_row)

        frame = pd.DataFrame(canonical_rows)
        if frame.empty:
            raise InputValidationError(
                "No valid registration rows remain after validation.",
                code="NO_VALID_REGISTRATION_ROWS",
                context={"issue_count": len(issues)},
            )
        duplicate_mask = frame["player_id"].duplicated(keep=False)
        if duplicate_mask.any() and self.config.duplicate_player_rule != "reject_generation":
            keep: Literal["first", "last"] = (
                "last" if self.config.duplicate_player_rule == "keep_latest" else "first"
            )
            ordered = frame.sort_values(
                ["registration_date", "_source_row_number"]
                if keep == "last"
                else ["_source_row_number"]
            )
            discarded = ordered[ordered.duplicated("player_id", keep=keep)]
            for _, discarded_row in discarded.iterrows():
                issues.append(
                    ValidationRecord(
                        source_filename=workbook_path.name,
                        worksheet=worksheet_name,
                        source_row_number=int(discarded_row["_source_row_number"]),
                        record_identifier=str(discarded_row["player_id"]),
                        reason_code="DUPLICATE_PLAYER_ID_EXCLUDED",
                        reason_message=(
                            "Duplicate Player ID excluded by configured rule "
                            f"'{self.config.duplicate_player_rule}'."
                        ),
                        processing_stage="business_validation",
                        created_timestamp=now,
                    )
                )
        frame = self._resolve_duplicates(frame)
        return ValidatedWorkbook(
            worksheet=worksheet_name,
            source_filename=workbook_path.name,
            frame=frame,
            issues=issues,
            source_columns=source_columns,
        )

    def _resolve_columns(self, headers: list[str]) -> dict[str, str]:
        lookup = {_normalise_header(header): header for header in headers}
        resolved: dict[str, str] = {}
        for canonical, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                if _normalise_header(alias) in lookup:
                    resolved[canonical] = lookup[_normalise_header(alias)]
                    break
        return resolved

    def _resolve_duplicates(self, frame: pd.DataFrame) -> pd.DataFrame:
        duplicate_mask = frame["player_id"].duplicated(keep=False)
        if not duplicate_mask.any():
            return frame
        duplicate_ids = sorted(frame.loc[duplicate_mask, "player_id"].astype(str).unique().tolist())
        if self.config.duplicate_player_rule == "reject_generation":
            raise InputValidationError(
                "Duplicate Player IDs were found.",
                code="DUPLICATE_PLAYER_IDS",
                context={"player_ids": duplicate_ids[:100], "count": len(duplicate_ids)},
            )
        if self.config.duplicate_player_rule == "keep_latest":
            return (
                frame.sort_values(["registration_date", "_source_row_number"])
                .drop_duplicates("player_id", keep="last")
                .reset_index(drop=True)
            )
        return (
            frame.sort_values("_source_row_number")
            .drop_duplicates("player_id", keep="first")
            .reset_index(drop=True)
        )

    @staticmethod
    def _issue(
        path: Path,
        worksheet: str,
        row: dict[str, Any],
        identifier: str | None,
        code: str,
        message: str,
        created: datetime,
    ) -> ValidationRecord:
        return ValidationRecord(
            source_filename=path.name,
            worksheet=worksheet,
            source_row_number=int(row["_source_row_number"]),
            record_identifier=identifier,
            reason_code=code,
            reason_message=message,
            processing_stage="business_validation",
            created_timestamp=created,
        )
