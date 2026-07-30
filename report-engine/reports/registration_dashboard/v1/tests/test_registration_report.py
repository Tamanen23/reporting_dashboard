import json
from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from core.exceptions import InputValidationError
from reports.registration_dashboard.v1.config import RegistrationConfig
from reports.registration_dashboard.v1.report import RegistrationDashboardReport

HEADERS = [
    "ID",
    "User",
    "Registered Date",
    "Reg. finished",
    "Status",
    "Disabled",
    "Deleted",
    "Last Deposit",
    "Email",
    "Mobile Number",
    "Country",
]


def create_valid_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "User List-28"
    sheet.append(HEADERS)
    sheet.append(
        [
            "P001",
            "alpha",
            date(2026, 7, 20),
            "Completed",
            "active",
            "no",
            "no",
            date(2026, 7, 21),
            "a@example.test",
            "111",
            "MU",
        ]
    )
    sheet.append(
        [
            "P002",
            "bravo",
            date(2026, 7, 20),
            "No",
            "pending",
            "no",
            "no",
            None,
            "b@example.test",
            "222",
            "MU",
        ]
    )
    sheet.append(
        [
            "P003",
            "charlie",
            date(2026, 7, 21),
            "Verified",
            "active",
            "no",
            "no",
            date(2026, 7, 22),
            "c@example.test",
            "333",
            "MU",
        ]
    )
    sheet.append(
        [
            "P004",
            "delta",
            date(2026, 7, 23),
            "No",
            "active",
            "yes",
            "no",
            None,
            "d@example.test",
            "444",
            "MU",
        ]
    )
    sheet.append(
        [
            "P005",
            "TestUser",
            date(2026, 7, 20),
            "Completed",
            "active",
            "no",
            "no",
            None,
            "",
            "",
            "MU",
        ]
    )
    sheet.append(
        [None, "blank-id", date(2026, 7, 20), "Completed", "active", "no", "no", None, "", "", "MU"]
    )
    sheet.append(
        [
            "P006",
            "invalid-date",
            "not-a-date",
            "Completed",
            "active",
            "no",
            "no",
            None,
            "",
            "",
            "MU",
        ]
    )
    sheet.append(
        [
            "P007",
            "deleted-user",
            date(2026, 7, 22),
            "Completed",
            "active",
            "no",
            "yes",
            None,
            "",
            "",
            "MU",
        ]
    )
    sheet.append(
        [
            "P008",
            "report-date-user",
            date(2026, 7, 24),
            "Completed",
            "active",
            "no",
            "no",
            None,
            "",
            "",
            "MU",
        ]
    )
    workbook.save(path)
    return path


def run_report(workbook: Path, output: Path, render: bool = False) -> dict[str, Path]:
    return RegistrationDashboardReport().run(
        workbook,
        output,
        report_date=date(2026, 7, 24),
        reporting_period_start=date(2026, 7, 20),
        reporting_period_end=date(2026, 7, 24),
        generation_uuid="00000000-0000-4000-8000-000000000001",
        render_outputs=render,
    )


def test_pipeline_generates_exact_results_and_traceable_artifacts(tmp_path: Path) -> None:
    artifacts = run_report(create_valid_workbook(tmp_path / "users.xlsx"), tmp_path / "run")
    result = json.loads(artifacts["calculated_results"].read_text())
    expected = json.loads((Path(__file__).parent / "expected-results.json").read_text())

    assert result["summary"] == expected["summary"]
    assert result["rates"] == expected["rates"]
    assert result["averages"] == {
        "registrations_per_day": "1.00",
        "completed_registrations_per_day": "0.60",
        "registered_and_deposited_per_day": "0.40",
        "average_ftd_per_day": "0.40",
    }
    assert result["highest_registration_day"] == {"date": "2026-07-20", "value": 2}
    assert all(check["passed"] for check in result["reconciliation"])
    assert result["last_ten_days_total"] == 5
    assert len(artifacts["registration_dataset"].read_bytes()) > 100
    issue_codes = {
        issue["reason_code"] for issue in json.loads(artifacts["validation_log"].read_text())
    }
    assert issue_codes == {
        "TEST_ACCOUNT",
            "BLANK_PLAYER_ID",
            "INVALID_REGISTRATION_DATE",
            "DELETED_ACCOUNT",
        }
    manifest = json.loads(artifacts["manifest"].read_text())
    assert manifest["report_code"] == "registration_dashboard"
    assert len(manifest["inputs"][0]["sha256"]) == 64
    assert {artifact["key"] for artifact in manifest["artifacts"]} >= {
        "registration_dataset",
        "validation_log",
        "calculated_results",
        "reconciliation_report",
        "chart_funnel",
        "chart_last_ten_days",
    }


def test_registration_csv_generates_results(tmp_path: Path) -> None:
    csv_path = tmp_path / "users.csv"
    pd.DataFrame([{
        "ID": "P001", "User": "alpha", "Registered At": "06/07/26 13:32:32",
        "Reg. finished": "Completed", "Status": "active", "Disabled": "no",
        "Deleted": "no", "Last Deposit": "07/07/26 10:15:00",
    }]).to_csv(csv_path, index=False)
    artifacts = RegistrationDashboardReport().run(
        csv_path, tmp_path / "csv-run",
        report_date=date(2026, 7, 10),
        reporting_period_start=date(2026, 7, 5),
        reporting_period_end=date(2026, 7, 8),
        generation_uuid="csv-date-test",
        render_outputs=False,
    )
    result = json.loads(artifacts["calculated_results"].read_text())
    assert result["summary"]["total_registrations"] == 1
    assert result["summary"]["completed_registrations"] == 1
    assert result["highest_registration_day"] == {"date": "2026-07-06", "value": 1}
    assert "EXCLUDING" not in artifacts["dashboard_html"].read_text()


def test_duplicate_player_ids_fail_under_provisional_strict_default(tmp_path: Path) -> None:
    path = create_valid_workbook(tmp_path / "duplicates.xlsx")
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    workbook["User List-28"].append(
        [
            "P001",
            "duplicate",
            date(2026, 7, 22),
            "Completed",
            "active",
            "no",
            "no",
            None,
            "",
            "",
            "MU",
        ]
    )
    workbook.save(path)
    with pytest.raises(InputValidationError) as error:
        run_report(path, tmp_path / "run")
    assert error.value.code == "DUPLICATE_PLAYER_IDS"


def test_configured_duplicate_resolution_logs_discarded_row(tmp_path: Path) -> None:
    path = create_valid_workbook(tmp_path / "duplicates.xlsx")
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    workbook["User List-28"].append(
        [
            "P001",
            "replacement",
            date(2026, 7, 22),
            "Completed",
            "active",
            "no",
            "no",
            None,
            "",
            "",
            "MU",
        ]
    )
    workbook.save(path)
    report = RegistrationDashboardReport(RegistrationConfig(duplicate_player_rule="keep_latest"))
    artifacts = report.run(
        path,
        tmp_path / "run",
        report_date=date(2026, 7, 24),
        reporting_period_start=date(2026, 7, 20),
        reporting_period_end=date(2026, 7, 24),
        generation_uuid="00000000-0000-4000-8000-000000000002",
        render_outputs=False,
    )
    issues = json.loads(artifacts["validation_log"].read_text())
    assert any(issue["reason_code"] == "DUPLICATE_PLAYER_ID_EXCLUDED" for issue in issues)


@pytest.mark.render
def test_end_to_end_generates_verified_pdf_and_png(tmp_path: Path) -> None:
    artifacts = run_report(
        create_valid_workbook(tmp_path / "users.xlsx"), tmp_path / "rendered", render=True
    )
    assert artifacts["pdf"].read_bytes().startswith(b"%PDF")
    assert artifacts["png"].read_bytes().startswith(b"\x89PNG")
