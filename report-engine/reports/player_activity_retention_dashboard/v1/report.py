from __future__ import annotations

import base64
import hashlib
import json
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import pandas as pd
from jinja2 import Environment, FileSystemLoader, StrictUndefined
from openpyxl import load_workbook
from PIL import Image
from playwright.sync_api import sync_playwright
from pypdf import PdfReader

from core.contracts import BaseReport
from core.exceptions import InputValidationError
from core.reporting_period import validate_reporting_period
from core.tabular import parse_datetime, parse_numeric, read_table

from .config import PlayerActivityConfig

VERSION = "1.0.0-provisional.6"
USER_HEADERS = ["ID", "User", "Registered Date", "Reg. finished", "Disabled", "Deleted"]
PAYMENT_HEADERS = ["Username", "User ID", "Amount", "Gateway", "Processed", "Type", "Processed Date", "Status"]
BET_LEGS_HEADERS = ["Slip #", "User #", "User Name", "Issue Time", "Slip State", "Bet Status", "Game", "Stake"]
CSV_HEADER_ALIASES = {
    "Processed at": "Processed Date",
    "Registered At": "Registered Date",
}


class PlayerActivityRetentionDashboardReport(BaseReport):
    def __init__(self, config: PlayerActivityConfig | None = None):
        self.config = config or PlayerActivityConfig()

    def run(
        self,
        input_paths: dict[str, Path],
        work_directory: Path,
        *,
        report_date: date,
        reporting_period_start: date,
        reporting_period_end: date,
        generation_uuid: str,
        render_outputs: bool = True,
    ) -> dict[str, Path]:
        validate_reporting_period(
            report_date, reporting_period_start, reporting_period_end
        )
        required = {"user_list", "payment_transactions", "bet_legs"}
        if set(input_paths) != required:
            raise InputValidationError(
                "Player Activity requires User List, Deposits & Withdrawals, and Bet Legs workbooks.",
                code="PLAYER_ACTIVITY_INPUTS_MISSING",
                context={"missing": sorted(required - set(input_paths))},
            )
        work_directory = Path(work_directory)
        for name in ("prepared", "results", "charts", "render", "outputs", "manifest"):
            (work_directory / name).mkdir(parents=True, exist_ok=True)

        effective_end = reporting_period_end
        users, user_issues = self._read_users(
            input_paths["user_list"], reporting_period_start, effective_end
        )
        payments, payment_issues = self._read_payments(input_paths["payment_transactions"], effective_end)
        bets, bet_issues = self._read_bets(input_paths["bet_legs"], effective_end)
        source_coverage = self._source_coverage(users, payments, bets, effective_end)
        master = self._master(users, payments, bets, reporting_period_start, effective_end)
        results = self._calculate(master, bets, report_date, reporting_period_start, reporting_period_end)
        results["source_coverage"] = source_coverage
        results["warnings"].extend(
            item["message"] for item in source_coverage.values() if item["status"] == "review"
        )

        master_path = work_directory / "prepared" / "master-player-dataset.parquet"
        master.to_parquet(master_path, index=False)
        crm_path = work_directory / "prepared" / "player-classification-and-crm-targets.csv"
        crm_columns = [
            "player_id",
            "username",
            "player_classification",
            "active_last_7_days",
            "regular_player_5_plus_days",
            "highly_engaged_10_plus_days",
            "vip_player",
            "crm_target",
            "priority_crm_target",
            "crm_target_reason",
            "activity_segment",
            "value_segment",
            "distinct_betting_days",
            "bet_count",
            "stake",
            "last_bet_date",
            "deposit_count",
            "deposit_amount",
            "last_deposit_date",
            "last_activity_date",
            "excluded_from_kpis",
        ]
        crm_export = master[crm_columns].copy()
        for column in (
            "active_last_7_days",
            "regular_player_5_plus_days",
            "highly_engaged_10_plus_days",
            "vip_player",
            "crm_target",
            "priority_crm_target",
            "excluded_from_kpis",
        ):
            crm_export[column] = crm_export[column].map({True: "Yes", False: "No"})
        crm_export.to_csv(crm_path, index=False, date_format="%Y-%m-%d")
        validation = {
            "report_code": "player_activity_retention_dashboard",
            "source_rows": {
                "registration_players": len(users), "payments": len(payments), "settled_bet_slips": len(bets),
            },
            "issues": user_issues + payment_issues + bet_issues,
            "source_coverage": source_coverage,
            "excluded_player_count": int(master["excluded"].sum()),
            "provisional_rules": [
                "Settled Bet Legs rows are the authoritative sports and casino betting source.",
                "Dormancy is measured against the effective reporting-period end.",
                "VIP is the top configured percentile by lifetime deposit amount.",
            ],
            "source_discrepancy": (
                {
                    "reference_betting_players": self.config.audit_reference_betting_players,
                    "bet_legs_betting_players": int(bets.player_id.nunique()),
                    "message": "The supplied production reference and source workbook do not contain the same betting-player population.",
                }
                if reporting_period_start == self.config.audit_reference_period_start
                and reporting_period_end == self.config.audit_reference_period_end
                else None
            ),
        }
        validation_path = work_directory / "prepared" / "validation-log.json"
        validation_path.write_text(json.dumps(validation, indent=2, default=str), encoding="utf-8")
        results_path = work_directory / "results" / "calculated-results.json"
        results_path.write_text(json.dumps(results, indent=2), encoding="utf-8")
        reconciliation_path = work_directory / "results" / "reconciliation-report.json"
        reconciliation_path.write_text(json.dumps(results["reconciliation_report"], indent=2), encoding="utf-8")

        chart_path = self._chart(results, work_directory / "charts")
        html_path = self._html(results, chart_path, work_directory / "render")
        artifacts = {
            "master_player_dataset": master_path,
            "crm_segment_export": crm_path,
            "validation_log": validation_path,
            "calculated_results": results_path,
            "reconciliation_report": reconciliation_path,
            "chart_player_segments": chart_path,
            "dashboard_html": html_path,
        }
        if render_outputs:
            pdf_path, png_path = self._outputs(html_path, work_directory / "outputs")
            self._verify(results, html_path, pdf_path, png_path)
            artifacts.update(pdf=pdf_path, png=png_path)

        manifest_path = work_directory / "manifest" / "manifest.json"
        manifest = {
            "generation_uuid": generation_uuid,
            "report_code": "player_activity_retention_dashboard",
            "report_date": report_date.isoformat(),
            "reporting_period_start": reporting_period_start.isoformat(),
            "reporting_period_end": reporting_period_end.isoformat(),
            "effective_period_end": effective_end.isoformat(),
            "timezone": self.config.timezone,
            "definition_version": VERSION,
            "calculation_version": VERSION,
            "template_version": VERSION,
            "generated_at": datetime.now(UTC).isoformat(),
            "inputs": [
                {"key": key, "filename": path.name, "sha256": self._sha(path)}
                for key, path in input_paths.items()
            ],
            "source_mapping": {
                "registration": {"sheet": self.config.user_worksheet, "player_id": "ID", "username": "User", "registration_date": "Registered Date", "completed": "Reg. finished", "disabled": "Disabled", "deleted": "Deleted"},
                "payments": {"sheet": self.config.payment_worksheet, "player_id": "User ID", "username": "Username", "amount": "Amount", "gateway": "Gateway", "type": "Type", "date": "Processed Date", "status": "Status"},
                "betting": {"sheet": self.config.bet_legs_worksheet, "bet_id": "Slip #", "player_id": "User #", "username": "User Name", "date": "Issue Time", "game": "Game", "slip_state": "Slip State", "bet_status": "Bet Status", "stake": "Stake"},
            },
            "configuration": {
                "betting_source": self.config.betting_source,
                "settled_bet_statuses": sorted(self.config.settled_bet_statuses),
                "successful_payment_statuses": sorted(self.config.successful_statuses),
                "allowed_payment_gateways": sorted(self.config.allowed_gateways),
                "exclude_test_usernames": True,
                "date_normalization": "All source timestamps are parsed and normalized to calendar dates.",
                "dormancy_days": self.config.dormancy_days,
                "exclude_disabled_accounts": self.config.exclude_disabled_accounts,
                "exclude_deleted_accounts": self.config.exclude_deleted_accounts,
                "vip_percentile": self.config.vip_percentile,
                "value_basis": self.config.value_basis,
                "provisional": True,
            },
            "artifacts": [
                {"key": key, "filename": path.name, "relative_path": str(path.relative_to(work_directory)), "size_bytes": path.stat().st_size, "sha256": self._sha(path)}
                for key, path in artifacts.items()
            ],
            "warnings": results["warnings"],
        }
        manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
        artifacts["manifest"] = manifest_path
        return artifacts

    def _source_coverage(
        self,
        users: pd.DataFrame,
        payments: pd.DataFrame,
        bets: pd.DataFrame,
        effective_end: date,
    ) -> dict[str, dict[str, Any]]:
        sources = {
            "user_list": users["registration_date"],
            "payment_transactions": payments["transaction_date"],
            "bet_legs": bets["bet_date"],
        }
        coverage: dict[str, dict[str, Any]] = {}
        for key, values in sources.items():
            parsed = pd.to_datetime(values, errors="coerce").dropna()
            minimum = parsed.min().date() if not parsed.empty else None
            maximum = parsed.max().date() if not parsed.empty else None
            is_current = maximum is not None and maximum >= effective_end
            coverage[key] = {
                "minimum_date": minimum.isoformat() if minimum else None,
                "maximum_date": maximum.isoformat() if maximum else None,
                "requested_effective_end": effective_end.isoformat(),
                "status": "current" if is_current else "review",
                "message": (
                    f"{key.replace('_', ' ').title()} contains records only through "
                    f"{maximum.isoformat() if maximum else 'an unknown date'}; "
                    f"the requested effective end is {effective_end.isoformat()}."
                ),
            }
        return coverage

    def _headers(self, path: Path, sheet: str, required: list[str]) -> None:
        if path.suffix.casefold() == ".csv":
            headers = [
                CSV_HEADER_ALIASES.get(str(value).strip(), str(value).strip())
                for value in read_table(path).columns
            ]
            missing = [value for value in required if value not in headers]
            if missing:
                raise InputValidationError("A Player Activity CSV file has missing columns.", code="HEADERS_INVALID", context={"missing": missing})
            return
        workbook = load_workbook(path, read_only=True, data_only=True)
        if sheet not in workbook.sheetnames:
            raise InputValidationError(f"Required worksheet '{sheet}' was not found.", code="WORKSHEET_MISSING")
        observed = [str(cell.value or "").strip() for cell in next(workbook[sheet].iter_rows(min_row=1, max_row=1))]
        workbook.close()
        missing = [name for name in required if name not in observed]
        if missing:
            raise InputValidationError("A Player Activity source workbook has missing columns.", code="HEADERS_INVALID", context={"sheet": sheet, "missing": missing})

    def _read_users(
        self, path: Path, effective_start: date, effective_end: date
    ) -> tuple[pd.DataFrame, list[dict]]:
        self._headers(path, self.config.user_worksheet, USER_HEADERS)
        frame = read_table(path, sheet_name=self.config.user_worksheet)
        frame = frame.rename(columns={"ID": "player_id", "User": "username", "Registered Date": "registration_date", "Registered At": "registration_date", "Reg. finished": "registration_completed", "Disabled": "disabled", "Deleted": "deleted"})
        issues = []
        blank = frame.player_id.isna() | frame.player_id.astype(str).str.strip().str.casefold().isin({"", "nan", "none", "null"})
        for row in frame.index[blank]: issues.append({"source": "user_list", "row": int(row) + 2, "code": "BLANK_PLAYER_ID"})
        frame = frame[~blank].copy()
        frame["player_id"] = frame.player_id.astype(str).str.replace(r"\.0$", "", regex=True)
        duplicates = frame.player_id.duplicated(keep=False)
        if duplicates.any():
            raise InputValidationError("Duplicate Player IDs are not allowed in the User List.", code="DUPLICATE_PLAYER_ID", context={"count": int(duplicates.sum())})
        frame["registration_date"] = parse_datetime(
            frame.registration_date, csv_source=path.suffix.casefold() == ".csv"
        ).dt.normalize()
        in_period = frame.registration_date.dt.date.between(effective_start, effective_end)
        if (~in_period).any():
            issues.append({
                "source": "user_list",
                "code": "OUTSIDE_REPORTING_PERIOD",
                "count": int((~in_period).sum()),
            })
        frame = frame[in_period].copy()
        frame["completed"] = frame.registration_completed.astype(str).str.strip().str.casefold().isin(self.config.completed_values)
        frame["disabled_flag"] = frame.disabled.astype(str).str.strip().str.casefold().eq("yes")
        frame["deleted_flag"] = frame.deleted.astype(str).str.strip().str.casefold().eq("yes")
        frame["test_flag"] = frame.username.astype(str).str.contains("test", case=False, na=False)
        if frame["test_flag"].any():
            issues.append({"source": "user_list", "code": "TEST_ACCOUNTS_EXCLUDED", "count": int(frame["test_flag"].sum())})
            frame = frame[~frame["test_flag"]].copy()
        return frame[["player_id", "username", "registration_date", "completed", "disabled_flag", "deleted_flag", "test_flag"]], issues

    def _read_users_dataset(
        self, path: Path, effective_start: date, effective_end: date
    ) -> tuple[pd.DataFrame, list[dict]]:
        frame = pd.read_parquet(path)
        required = {"player_id", "username", "registration_date", "registration_completed", "is_disabled", "is_deleted"}
        missing = sorted(required - set(frame.columns))
        if missing:
            raise InputValidationError(
                "The Registration prepared dataset is incompatible with Player Activity.",
                code="REGISTRATION_DATASET_INCOMPATIBLE",
                context={"missing_columns": missing},
            )
        frame = frame.rename(columns={
            "registration_completed": "completed",
            "is_disabled": "disabled_flag",
            "is_deleted": "deleted_flag",
        }).copy()
        frame["player_id"] = frame.player_id.astype(str).str.replace(r"\.0$", "", regex=True)
        frame["registration_date"] = pd.to_datetime(frame.registration_date, errors="coerce").dt.normalize()
        in_period = frame.registration_date.dt.date.between(effective_start, effective_end)
        issues = []
        if (~in_period).any():
            issues.append({
                "source": "registration_dataset",
                "code": "OUTSIDE_REPORTING_PERIOD",
                "count": int((~in_period).sum()),
            })
        frame = frame[in_period].copy()
        frame["test_flag"] = frame.username.astype(str).str.contains("test", case=False, na=False)
        if frame["test_flag"].any():
            issues.append({"source": "registration_dataset", "code": "TEST_ACCOUNTS_EXCLUDED", "count": int(frame["test_flag"].sum())})
            frame = frame[~frame["test_flag"]].copy()
        duplicates = frame.player_id.duplicated(keep=False)
        if duplicates.any():
            raise InputValidationError(
                "The Registration prepared dataset contains duplicate Player IDs.",
                code="DUPLICATE_PLAYER_ID",
                context={"count": int(duplicates.sum())},
            )
        return frame[["player_id", "username", "registration_date", "completed", "disabled_flag", "deleted_flag", "test_flag"]], issues

    def _read_payments(self, path: Path, effective_end: date) -> tuple[pd.DataFrame, list[dict]]:
        self._headers(path, self.config.payment_worksheet, PAYMENT_HEADERS)
        frame = read_table(path, sheet_name=self.config.payment_worksheet)
        frame = frame.rename(columns={"User ID": "player_id", "Username": "username", "Amount": "amount", "Gateway": "gateway", "Type": "transaction_type", "Processed Date": "transaction_date", "Processed at": "transaction_date", "Status": "status", "Processed": "processed"})
        frame["player_id"] = frame.player_id.astype(str).str.replace(r"\.0$", "", regex=True)
        frame["transaction_date"] = parse_datetime(
            frame.transaction_date, csv_source=path.suffix.casefold() == ".csv"
        ).dt.normalize()
        numeric_amount = parse_numeric(frame.amount)
        valid_player_id = ~frame.player_id.str.strip().str.casefold().isin({"", "nan", "none", "null"})
        valid = (
            valid_player_id
            & frame.status.astype(str).str.strip().str.casefold().isin(self.config.successful_statuses)
            & frame.processed.astype(str).str.strip().str.casefold().eq("yes")
            & frame.gateway.astype(str).str.strip().str.casefold().isin(self.config.allowed_gateways)
            & ~frame.username.astype(str).str.contains("test", case=False, na=False)
            & frame.transaction_date.notna()
            & numeric_amount.notna()
            & (frame.transaction_date.dt.date <= effective_end)
            & ~frame.transaction_date.dt.date.isin(self.config.excluded_dates)
        )
        issues = [{"source": "payments", "code": "REJECTED_PAYMENT_ROWS", "count": int((~valid).sum())}] if (~valid).any() else []
        frame = frame[valid].copy()
        frame["transaction_type"] = frame.transaction_type.astype(str).str.strip().str.casefold()
        frame["amount"] = numeric_amount[valid]
        return frame[["player_id", "username", "amount", "transaction_type", "transaction_date"]], issues

    def _read_payments_dataset(self, path: Path, effective_end: date) -> tuple[pd.DataFrame, list[dict]]:
        frame = pd.read_parquet(path)
        required = {"user_id", "username", "amount", "gateway", "transaction_type", "transaction_date"}
        missing = sorted(required - set(frame.columns))
        if missing:
            raise InputValidationError(
                "The Payments prepared dataset is incompatible with Player Activity.",
                code="PAYMENT_DATASET_INCOMPATIBLE",
                context={"missing_columns": missing},
            )
        frame = frame.rename(columns={"user_id": "player_id"}).copy()
        frame["player_id"] = frame.player_id.astype(str).str.replace(r"\.0$", "", regex=True)
        frame["transaction_date"] = pd.to_datetime(frame.transaction_date, errors="coerce").dt.normalize()
        valid = (
            frame.transaction_date.notna()
            & (frame.transaction_date.dt.date <= effective_end)
            & ~frame.transaction_date.dt.date.isin(self.config.excluded_dates)
            & frame.gateway.astype(str).str.strip().str.casefold().isin(self.config.allowed_gateways)
            & ~frame.username.astype(str).str.contains("test", case=False, na=False)
        )
        issues = [{"source": "payment_dataset", "code": "OUT_OF_SCOPE_PAYMENT_ROWS", "count": int((~valid).sum())}] if (~valid).any() else []
        frame = frame[valid].copy()
        frame["amount"] = parse_numeric(frame.amount).fillna(0)
        return frame[["player_id", "username", "amount", "transaction_type", "transaction_date"]], issues

    def _read_bets(self, path: Path, effective_end: date) -> tuple[pd.DataFrame, list[dict]]:
        self._headers(path, self.config.bet_legs_worksheet, BET_LEGS_HEADERS)
        frame = read_table(path, sheet_name=self.config.bet_legs_worksheet)
        frame = frame.rename(columns={"Slip #": "bet_id", "User #": "player_id", "User Name": "username", "Issue Time": "bet_date", "Slip State": "slip_state", "Bet Status": "bet_status", "Game": "game", "Stake": "stake"})
        frame["player_id"] = frame.player_id.astype(str).str.replace(r"\.0$", "", regex=True)
        frame["bet_date"] = parse_datetime(
            frame.bet_date, csv_source=path.suffix.casefold() == ".csv"
        )
        numeric_stake = parse_numeric(frame.stake)
        valid = (
            frame.bet_status.astype(str).str.strip().str.casefold().isin(self.config.settled_bet_statuses)
            & ~frame.player_id.str.strip().str.casefold().isin({"", "nan", "none", "null"})
            & ~frame.username.astype(str).str.contains("test", case=False, na=False)
            & frame.bet_date.notna()
            & numeric_stake.notna()
            & (frame.bet_date.dt.date <= effective_end)
            & ~frame.bet_date.dt.date.isin(self.config.excluded_dates)
        )
        issues = [{"source": "bet_legs", "code": "UNSETTLED_OR_CANCELLED_LEGS_EXCLUDED", "count": int((~valid).sum())}]
        frame = frame[valid].copy()
        frame["stake"] = numeric_stake[valid].abs()
        # A sportsbook combination produces multiple leg rows for one slip. Player activity,
        # stake and playing-day measures operate at unique-slip grain.
        frame = (
            frame.sort_values(["bet_id", "bet_date"])
            .groupby("bet_id", as_index=False)
            .agg(
                player_id=("player_id", "first"),
                username=("username", "first"),
                bet_date=("bet_date", "first"),
                stake=("stake", "max"),
                game=("game", lambda values: "Sports" if values.astype(str).str.casefold().eq("sports").any() else values.iloc[0]),
                bet_status=("bet_status", "first"),
            )
        )
        return frame[["bet_id", "player_id", "username", "bet_date", "stake", "game", "bet_status"]], issues

    def _master(self, users: pd.DataFrame, payments: pd.DataFrame, bets: pd.DataFrame, start: date, end: date) -> pd.DataFrame:
        deposits = payments[payments.transaction_type.eq("deposit")]
        withdrawals = payments[payments.transaction_type.eq("withdrawal")]
        dep = deposits.groupby("player_id").agg(first_deposit_date=("transaction_date", "min"), last_deposit_date=("transaction_date", "max"), deposit_count=("amount", "size"), deposit_amount=("amount", "sum"))
        wit = withdrawals.groupby("player_id").agg(withdrawal_count=("amount", "size"), withdrawal_amount=("amount", "sum"))
        bet = bets.groupby("player_id").agg(first_bet_date=("bet_date", "min"), last_bet_date=("bet_date", "max"), bet_count=("bet_id", "nunique"), stake=("stake", "sum"), distinct_betting_days=("bet_date", lambda values: values.dt.normalize().nunique()))
        master = users.set_index("player_id").join(dep).join(wit).join(bet).reset_index()
        for column in ("deposit_count", "deposit_amount", "withdrawal_count", "withdrawal_amount", "bet_count", "stake", "distinct_betting_days"):
            master[column] = master[column].fillna(0)
        master["excluded"] = master.test_flag | (master.disabled_flag & self.config.exclude_disabled_accounts) | (master.deleted_flag & self.config.exclude_deleted_accounts) | ~master.completed
        last_dates = pd.concat([master.last_deposit_date, master.last_bet_date], axis=1)
        master["last_activity_date"] = last_dates.max(axis=1)
        cutoff = pd.Timestamp(end - timedelta(days=self.config.dormancy_days))
        master["activity_segment"] = "registered_only"
        master.loc[master.deposit_count.gt(0) & master.bet_count.eq(0), "activity_segment"] = "deposited_never_bet"
        master.loc[master.distinct_betting_days.eq(1), "activity_segment"] = "one_time_player"
        master.loc[master.distinct_betting_days.between(2, 4), "activity_segment"] = "occasional_player"
        master.loc[master.distinct_betting_days.between(5, 9), "activity_segment"] = "regular_player"
        master.loc[master.distinct_betting_days.between(10, 19), "activity_segment"] = "highly_engaged"
        master.loc[master.distinct_betting_days.ge(20), "activity_segment"] = "core_player"
        master.loc[master.bet_count.gt(0) & master.last_bet_date.lt(cutoff), "activity_segment"] = "dormant_player"
        master.loc[master.excluded, "activity_segment"] = "excluded_or_incomplete"
        # Independent production-reference value bands. Ranking remains deterministic
        # when deposit totals tie by using Player ID as the secondary key.
        ranked = master.sort_values(["deposit_amount", "player_id"], ascending=[False, True])
        population = len(ranked)
        vip_end = round(population * .01)
        high_end = round(population * .05)
        regular_end = round(population * .20)
        standard_end = round(population * .50)
        master["value_segment"] = "low_value"
        master.loc[ranked.iloc[:vip_end].index, "value_segment"] = "vip"
        master.loc[ranked.iloc[vip_end:high_end].index, "value_segment"] = "high_value"
        master.loc[ranked.iloc[high_end:regular_end].index, "value_segment"] = "regular_value"
        master.loc[ranked.iloc[regular_end:standard_end].index, "value_segment"] = "standard"
        active_last_7_start = end - timedelta(days=6)
        active_last_7_ids = set(
            bets[bets.bet_date.dt.date.between(active_last_7_start, end)].player_id
        )
        master["active_last_7_days"] = master.player_id.isin(active_last_7_ids) & ~master.excluded
        master["regular_player_5_plus_days"] = master.distinct_betting_days.ge(5) & ~master.excluded
        master["highly_engaged_10_plus_days"] = master.distinct_betting_days.ge(10) & ~master.excluded
        master["vip_player"] = master.value_segment.eq("vip") & ~master.excluded
        master["crm_target"] = (
            master.activity_segment.isin(
                ["dormant_player", "deposited_never_bet", "one_time_player"]
            )
            & ~master.excluded
        )
        master["priority_crm_target"] = master.crm_target
        master["crm_target_reason"] = ""
        master.loc[
            master.activity_segment.eq("dormant_player") & master.crm_target,
            "crm_target_reason",
        ] = "Dormant player"
        master.loc[
            master.activity_segment.eq("deposited_never_bet") & master.crm_target,
            "crm_target_reason",
        ] = "Deposited but never bet"
        master.loc[
            master.activity_segment.eq("one_time_player") & master.crm_target,
            "crm_target_reason",
        ] = "One-time player"
        master["excluded_from_kpis"] = master.excluded

        classification_flags = [
            ("active_last_7_days", "Active"),
            ("regular_player_5_plus_days", "Regular"),
            ("highly_engaged_10_plus_days", "Highly Engaged"),
            ("vip_player", "VIP"),
        ]
        master["player_classification"] = master.apply(
            lambda row: " | ".join(
                label for column, label in classification_flags if bool(row[column])
            )
            or "Other",
            axis=1,
        )
        period_bets = bets[bets.bet_date.dt.date.between(start, end)]
        active_ids = set(period_bets.player_id)
        master["active_in_period"] = master.player_id.isin(active_ids)
        return master

    def _calculate(self, master: pd.DataFrame, bets: pd.DataFrame, report_date: date, start: date, period_end: date) -> dict[str, Any]:
        end = period_end
        valid = master[~master.excluded]
        segments = master.activity_segment.value_counts().to_dict()
        labels = {
            "registered_only": "Registered Only", "deposited_never_bet": "Deposited, Never Bet",
            "one_time_player": "One-Time Player", "occasional_player": "Occasional Player",
            "regular_player": "Regular Player", "highly_engaged": "Highly Engaged",
            "core_player": "Core Player", "dormant_player": "Dormant Player",
            "excluded_or_incomplete": "Excluded / Incomplete",
        }
        segment_rows = [{"key": key, "label": labels[key], "count": int(segments.get(key, 0)), "percentage": round(segments.get(key, 0) / len(master) * 100, 1) if len(master) else 0} for key in labels]
        active7 = int(master.active_last_7_days.sum())
        latest_bet_date = min(end, bets.bet_date.max().date()) if len(bets) else end
        yesterday_date = latest_bet_date - timedelta(days=1)
        today = set(bets[bets.bet_date.dt.date.eq(latest_bet_date)].player_id)
        yesterday = set(bets[bets.bet_date.dt.date.eq(yesterday_date)].player_id)
        played_both = len(today & yesterday)
        returning_rate = played_both / len(today) * 100 if today else 0
        playing = master[master.bet_count.gt(0) & ~master.excluded]
        avg_days = float(playing.distinct_betting_days.mean()) if len(playing) else 0
        vip = master[master.vip_player]
        gap_values = []
        for _, player_bets in bets.groupby("player_id"):
            dates = sorted(set(player_bets.bet_date.dt.normalize()))
            if len(dates) > 1:
                gap_values.extend((dates[index] - dates[index - 1]).days for index in range(1, len(dates)))
        average_gap = sum(gap_values) / len(gap_values) if gap_values else 0
        frequency = [
            {"label": "1 day", "count": int((playing.distinct_betting_days == 1).sum())},
            {"label": "2–4 days", "count": int(playing.distinct_betting_days.between(2, 4).sum())},
            {"label": "5–9 days", "count": int(playing.distinct_betting_days.between(5, 9).sum())},
            {"label": "10–19 days", "count": int(playing.distinct_betting_days.between(10, 19).sum())},
            {"label": "20+ days", "count": int((playing.distinct_betting_days >= 20).sum())},
        ]
        for item in frequency:
            item["percentage"] = round(item["count"] / len(master) * 100, 1) if len(master) else 0
        placed_first_bet = int((master.bet_count.gt(0) & ~master.excluded).sum())
        regular_players = int(master.regular_player_5_plus_days.sum())
        dormant_30 = int((master.bet_count.gt(0) & master.last_bet_date.lt(pd.Timestamp(end - timedelta(days=30))) & ~master.excluded).sum())
        dormant_60 = int((master.bet_count.gt(0) & master.last_bet_date.lt(pd.Timestamp(end - timedelta(days=60))) & ~master.excluded).sum())
        deposited_never_bet = int((master.deposit_count.gt(0) & master.bet_count.eq(0) & ~master.excluded).sum())
        one_time = int((master.distinct_betting_days.eq(1) & ~master.excluded).sum())
        priority_mask = master.crm_target
        priority_targets = int(priority_mask.sum())
        reconciliation = [
            {"name": "one_player_per_master_row", "passed": master.player_id.is_unique, "actual": len(master), "expected": master.player_id.nunique()},
            {"name": "activity_segments_equal_master_players", "passed": sum(row["count"] for row in segment_rows) == len(master), "actual": sum(row["count"] for row in segment_rows), "expected": len(master)},
            {"name": "excluded_test_accounts_not_in_crm_segments", "passed": not master[master.test_flag].activity_segment.ne("excluded_or_incomplete").any(), "actual": int(master.test_flag.sum()), "expected": int(master.test_flag.sum())},
            {"name": "active_players_not_above_valid_players", "passed": int(active7) <= len(master), "actual": int(active7), "expected": len(master)},
            {"name": "crm_export_flags_reconcile", "passed": int(master.crm_target.sum()) == priority_targets, "actual": int(master.crm_target.sum()), "expected": priority_targets},
        ]
        warnings = [{
            "code": "PROVISIONAL_PLAYER_ACTIVITY_RULES",
            "message": "Settled-bet statuses, dormancy and value thresholds require formal business approval.",
        }]
        if (
            start == self.config.audit_reference_period_start
            and end == self.config.audit_reference_period_end
            and bets.player_id.nunique()
            != self.config.audit_reference_betting_players
        ):
            warnings.append({
                "code": "PRODUCTION_REFERENCE_POPULATION_MISMATCH",
                "message": (
                    f"Bet Legs contains {bets.player_id.nunique()} settled betting players; "
                    "the supplied production reference shows "
                    f"{self.config.audit_reference_betting_players}."
                ),
            })
        if latest_bet_date < end:
            warnings.append({
                "code": "BET_LEGS_COVERAGE_INCOMPLETE",
                "message": f"The reporting period ends {end.isoformat()}, but Bet Legs ends {latest_bet_date.isoformat()}.",
            })
        if placed_first_bet > int(master.deposit_count.gt(0).sum()):
            warnings.append({
                "code": "BETTORS_EXCEED_DEPOSITORS",
                "message": (
                    f"{placed_first_bet} betting players exceed the "
                    f"{int(master.deposit_count.gt(0).sum())} players found in the compatible Payments dataset."
                ),
            })
        kpis = {
            "registered_players": len(master),
            "completed_registrations": int(master.completed.sum()),
            "depositors": int(master.deposit_count.gt(0).sum()),
            "active_players_last_7_days": int(active7),
            "placed_first_bet": placed_first_bet,
            "regular_players_5_plus_days": int(master.regular_player_5_plus_days.sum()),
            "highly_engaged_10_plus_days": int(master.highly_engaged_10_plus_days.sum()),
            "vip_players": len(vip),
            "valid_players": len(valid),
        }
        insights = [
            f"{kpis['depositors']:,} players ({kpis['depositors']/len(master)*100:.1f}% of registered) have made a successful deposit." if len(master) else "No registered players were found.",
            f"{kpis['regular_players_5_plus_days']:,} valid players have bet on five or more distinct days.",
            f"{kpis['active_players_last_7_days']:,} players were active during the final seven reporting days.",
            f"{segments.get('dormant_player', 0):,} players meet the configured {self.config.dormancy_days}-day dormancy rule.",
            f"{len(vip):,} players are provisionally classified as VIP by lifetime deposits.",
        ]
        return {
            "period_start": start.isoformat(),
            "period_end": end.isoformat(),
            "report": {
                "title": "Player Activity & Retention Dashboard",
                "report_date": report_date.strftime("%d %B %Y"),
                "period_start": start.strftime("%d %B %Y"),
                "period_end": end.strftime("%d %B %Y"),
                "excluded_dates": [
                    value.strftime("%d %B %Y")
                    for value in sorted(self.config.excluded_dates)
                ],
            },
            "kpis": kpis,
            "segments": segment_rows,
            "frequency": frequency,
            "funnel": [
                {"label": "Registered", "count": len(master), "rate": 100.0},
                {"label": "Completed Registration", "count": int(master.completed.sum()), "rate": round(int(master.completed.sum()) / len(master) * 100, 1) if len(master) else 0},
                {"label": "First Deposit", "count": int(master.deposit_count.gt(0).sum()), "rate": round(int(master.deposit_count.gt(0).sum()) / max(int(master.completed.sum()), 1) * 100, 1)},
                {"label": "Placed First Bet", "count": placed_first_bet, "rate": round(placed_first_bet / max(int(master.deposit_count.gt(0).sum()), 1) * 100, 1)},
                {"label": "Regular Players", "count": regular_players, "rate": round(regular_players / max(placed_first_bet, 1) * 100, 1)},
            ],
            "engagement": {"average_playing_days": round(avg_days, 1), "average_days_between_sessions": round(average_gap, 1), "average_bets": round(float(playing.bet_count.mean()) if len(playing) else 0, 1), "average_stake": round(float(playing.stake.mean()) if len(playing) else 0, 0), "average_deposit_count": round(float(master[master.deposit_count.gt(0)].deposit_count.mean()) if master.deposit_count.gt(0).any() else 0, 1), "average_deposit_amount": round(float(master[master.deposit_amount.gt(0)].deposit_amount.mean()) if master.deposit_amount.gt(0).any() else 0, 0)},
            "returning": {"yesterday_date": yesterday_date.isoformat(), "today_date": latest_bet_date.isoformat(), "source_through": latest_bet_date.isoformat(), "active_yesterday": len(yesterday), "active_today": len(today), "played_both": played_both, "return_rate": round(returning_rate, 1)},
            "reactivation": [
                {"label": "Dormant (30+ days)", "count": dormant_30},
                {"label": "Dormant (60+ days)", "count": dormant_60},
                {"label": "Deposited but never bet", "count": deposited_never_bet},
                {"label": "One-time players", "count": one_time},
                {"label": "Priority CRM targets", "count": priority_targets},
            ],
            "value_groups": [
                {
                    "key": label,
                    "label": {"vip": "VIP (Top 1%)", "high_value": "High Value (Next 4%)", "regular_value": "Regular (Next 15%)", "standard": "Standard (Next 30%)", "low_value": "Low Value (Bottom 50%)"}[label],
                    "count": int((master.value_segment == label).sum()),
                    "percentage": round(int((master.value_segment == label).sum()) / len(master) * 100, 1) if len(master) else 0,
                    "deposit_amount": round(float(master.loc[master.value_segment == label, "deposit_amount"].sum()), 0),
                    "deposit_percentage": round(float(master.loc[master.value_segment == label, "deposit_amount"].sum()) / max(float(master.deposit_amount.sum()), 1) * 100, 1),
                }
                for label in ("vip", "high_value", "regular_value", "standard", "low_value")
            ],
            "insights": insights,
            "warnings": warnings,
            "reconciliation_report": {"report_code": "player_activity_retention_dashboard", "passed": all(item["passed"] for item in reconciliation), "checks": reconciliation},
        }

    def _chart(self, results: dict, directory: Path) -> Path:
        directory.mkdir(parents=True, exist_ok=True)
        path = directory / "player-segments.png"
        plt.style.use("dark_background")
        fig, axis = plt.subplots(figsize=(4.2, 4.2), facecolor="#050505")
        rows = [row for row in results["segments"] if row["count"]]
        colors = ["#d81927", "#f17816", "#f2b705", "#4db414", "#1488d4", "#7541b6", "#999999", "#555555", "#b22c2c"]
        axis.pie(
            [row["count"] for row in rows],
            colors=colors[:len(rows)],
            startangle=90,
            autopct=lambda value: f"{value:.1f}%" if value >= 2 else "",
            textprops={"color": "white", "fontsize": 8},
            wedgeprops={"width": .50, "edgecolor": "#dddddd", "linewidth": .6},
        )
        axis.set_aspect("equal")
        fig.tight_layout(pad=.1)
        fig.savefig(path, dpi=150, facecolor=fig.get_facecolor(), transparent=False)
        plt.close(fig)
        return path

    def _html(self, results: dict, chart: Path, directory: Path) -> Path:
        directory.mkdir(parents=True, exist_ok=True)
        template_dir = Path(__file__).parent / "templates"
        env = Environment(loader=FileSystemLoader(template_dir), autoescape=True, undefined=StrictUndefined)
        logo_path = Path(__file__).parents[2] / "registration_dashboard" / "v1" / "assets" / "Favicon.jpeg"
        context = {
            **results,
            "chart_uri": f"data:image/png;base64,{base64.b64encode(chart.read_bytes()).decode()}",
            "logo_uri": f"data:image/jpeg;base64,{base64.b64encode(logo_path.read_bytes()).decode()}",
        }
        path = directory / "dashboard.html"
        path.write_text(env.get_template("dashboard.html").render(**context), encoding="utf-8")
        return path

    def _outputs(self, html: Path, directory: Path) -> tuple[Path, Path]:
        directory.mkdir(parents=True, exist_ok=True)
        pdf, png = directory / "dashboard.pdf", directory / "dashboard.png"
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            page = browser.new_page(viewport={"width": self.config.output_width, "height": self.config.output_height})
            page.goto(html.resolve().as_uri(), wait_until="networkidle")
            page.wait_for_function("document.fonts.status === 'loaded'")
            page.pdf(path=str(pdf), width=f"{self.config.output_width}px", height=f"{self.config.output_height}px", print_background=True, margin={"top": "0", "right": "0", "bottom": "0", "left": "0"})
            page.screenshot(path=str(png), type="png", full_page=True)
            browser.close()
        return pdf, png

    def _verify(self, results: dict, html: Path, pdf: Path, png: Path) -> None:
        text = html.read_text(encoding="utf-8")
        for value in (results["kpis"]["registered_players"], results["kpis"]["depositors"]):
            if f"{value:,}" not in text:
                raise RuntimeError(f"Calculated value {value} is missing from the dashboard.")
        if len(PdfReader(pdf).pages) != 1:
            raise RuntimeError("Player Activity PDF must contain exactly one page.")
        with Image.open(png) as image:
            if image.size != (self.config.output_width, self.config.output_height):
                raise RuntimeError(f"Unexpected dashboard PNG dimensions: {image.size}.")

    @staticmethod
    def _sha(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def validate_inputs(self, files, reporting_context): return []
    def normalize_inputs(self, files, work_directory, reporting_context): return {}
    def calculate(self, normalized_files, reporting_context): return {}
    def validate_results(self, results, reporting_context): return []
    def generate_charts(self, results, output_directory, reporting_context): return {}
    def get_template_context(self, results, charts, reporting_context): return {}
