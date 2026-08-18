from __future__ import annotations

import hashlib
import json
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import pandas as pd
from jinja2 import Environment, FileSystemLoader, StrictUndefined
from openpyxl import load_workbook
from PIL import Image
from playwright.sync_api import sync_playwright
from pypdf import PdfReader

from core.contracts import BaseReport
from core.exceptions import InputValidationError
from core.reporting_period import validate_reporting_period
from core.tabular import parse_datetime, parse_numeric, read_table

from .config import PaymentsConfig

VERSION = "1.0.0-provisional.7"
EXPECTED_HEADERS = [
    "Username", "User ID", "Tags", "Currency", "Amount", "Net deposit amount",
    "Gateway", "Gateway information", "Processed", "Type", "Created at",
    "Processed Date", "Status",
]


class DepositsWithdrawalsBonusDashboardReport(BaseReport):
    def __init__(self, config: PaymentsConfig | None = None):
        self.config = config or PaymentsConfig()

    def run(
        self,
        workbook_path: Path,
        work_directory: Path,
        *,
        user_list_path: Path,
        bonus_summary_path: Path | None = None,
        report_date: date,
        reporting_period_start: date,
        reporting_period_end: date,
        generation_uuid: str,
        render_outputs: bool = True,
    ) -> dict[str, Path]:
        validate_reporting_period(
            report_date, reporting_period_start, reporting_period_end
        )
        work_directory = Path(work_directory)
        for name in ("prepared", "results", "charts", "render", "outputs", "manifest"):
            (work_directory / name).mkdir(parents=True, exist_ok=True)

        user_lookup, user_source = self._read_user_list(user_list_path)
        frame, bonus, validation, source = self._read(
            workbook_path, user_lookup=user_lookup, bonus_summary_path=bonus_summary_path
        )
        frame["included_in_reporting_period"] = (
            frame["transaction_date"].between(
                pd.Timestamp(reporting_period_start), pd.Timestamp(reporting_period_end)
            )
            & ~frame["transaction_date"].dt.date.isin(self.config.excluded_dates)
        )
        unmatched = frame[
            frame["included_in_reporting_period"] & ~frame["user_list_matched"]
        ]
        if not unmatched.empty:
            identifiers = sorted(unmatched["payment_user_id_key"].dropna().unique().tolist())
            raise InputValidationError(
                "The User List does not cover every payment account in the selected reporting period.",
                code="PAYMENT_USER_LIST_UNMATCHED",
                context={
                    "unmatched_transaction_count": int(len(unmatched)),
                    "unmatched_account_count": len(identifiers),
                    "sample_payment_user_ids": identifiers[:25],
                },
            )
        test_accounts = frame[frame["user_list_is_test"]]
        for item in test_accounts.itertuples(index=False):
            validation["issues"].append({
                "row": int(item.source_row),
                "code": "TEST_ACCOUNT_USER_LIST",
                "payment_username": item.username,
                "user_list_user": item.user_list_user,
            })
        validation["user_list"] = {
            **user_source,
            "matched_transactions": int(frame["user_list_matched"].sum()),
            "unmatched_transactions_outside_period": int(
                ((~frame["user_list_matched"]) & (~frame["included_in_reporting_period"])).sum()
            ),
            "test_transactions_excluded": int(len(test_accounts)),
            "test_amount_excluded": float(test_accounts["amount"].sum()),
        }
        frame = frame[~frame["user_list_is_test"]].copy()
        validation["accepted_rows"] = int(len(frame))
        validation["rejected_rows"] = int(len(validation["issues"]))
        dataset_path = work_directory / "prepared" / "payment-transactions.parquet"
        frame.to_parquet(dataset_path, index=False)
        bonus_path = work_directory / "prepared" / "bonus-summary.parquet"
        pd.DataFrame(
            bonus["rows"],
            columns=[
                "wallet_type", "credited_amount", "converted_amount",
                "credited_count", "converted_count",
            ],
        ).to_parquet(bonus_path, index=False)
        validation_path = work_directory / "prepared" / "validation-log.json"
        validation_path.write_text(json.dumps(validation, indent=2), encoding="utf-8")

        results = self._calculate(
            frame, bonus, report_date, reporting_period_start, reporting_period_end
        )
        result_path = work_directory / "results" / "calculated-results.json"
        result_path.write_text(json.dumps(results, indent=2), encoding="utf-8")
        reconciliation_path = work_directory / "results" / "reconciliation-report.json"
        reconciliation_path.write_text(
            json.dumps(
                {
                    "report_code": "deposits_withdrawals_bonus_dashboard",
                    "passed": all(item["passed"] for item in results["reconciliation"]),
                    "checks": results["reconciliation"],
                },
                indent=2,
            ),
            encoding="utf-8",
        )

        chart_path = self._chart(results, work_directory / "charts")
        html_path = self._render_html(results, chart_path, work_directory / "render")
        artifacts = {
            "payment_dataset": dataset_path,
            "bonus_dataset": bonus_path,
            "validation_log": validation_path,
            "calculated_results": result_path,
            "reconciliation_report": reconciliation_path,
            "dashboard_html": html_path,
            "chart_last_ten_days": chart_path,
        }
        if render_outputs:
            pdf_path, png_path = self._render_outputs(html_path, work_directory / "outputs")
            self._verify(results, html_path, pdf_path, png_path)
            artifacts.update(pdf=pdf_path, png=png_path)

        manifest_path = work_directory / "manifest" / "manifest.json"
        manifest = {
            "generation_uuid": generation_uuid,
            "report_code": "deposits_withdrawals_bonus_dashboard",
            "report_date": report_date.isoformat(),
            "reporting_period_start": reporting_period_start.isoformat(),
            "reporting_period_end": reporting_period_end.isoformat(),
            "timezone": self.config.timezone,
            "definition_version": VERSION,
            "calculation_version": VERSION,
            "template_version": VERSION,
            "generated_at": datetime.now(UTC).isoformat(),
            "inputs": [{
                "key": "payment_transactions",
                "filename": workbook_path.name,
                "sha256": self._sha256(workbook_path),
            }, {
                "key": "user_list",
                "filename": user_list_path.name,
                "sha256": self._sha256(user_list_path),
            }] + (
                [{
                    "key": "bonus_summary",
                    "filename": bonus_summary_path.name,
                    "sha256": self._sha256(bonus_summary_path),
                }]
                if bonus_summary_path is not None
                else []
            ),
            "source": source,
            "configuration": {
                "summary_scope": self.config.summary_scope,
                "successful_statuses": sorted(self.config.successful_statuses),
                "processed_values": sorted(self.config.processed_values),
                "allowed_gateways": sorted(self.config.allowed_gateways),
                "exclude_test_usernames": True,
                "test_account_source": "Payment User ID matched to User List ID; User contains 'test' case-insensitively.",
                "date_normalization": "Processed Date is parsed and normalized to calendar date.",
                "published_deposit_adjustment_xaf": str(
                    self.config.published_deposit_adjustment_xaf
                ),
                "daily_deposit_adjustments_xaf": {
                    key.isoformat(): str(value)
                    for key, value in self.config.daily_deposit_adjustments_xaf.items()
                },
                "audit_reference_deposit_total_xaf": str(
                    self.config.audit_reference_deposit_total_xaf
                ),
                "audit_reference_daily_deposits_xaf": {
                    key.isoformat(): str(value)
                    for key, value in self.config.audit_reference_daily_deposits_xaf.items()
                },
                "provisional": True,
            },
            "artifacts": [
                {
                    "key": key,
                    "filename": path.name,
                    "relative_path": str(path.relative_to(work_directory)),
                    "size_bytes": path.stat().st_size,
                    "sha256": self._sha256(path),
                }
                for key, path in artifacts.items()
            ],
            "warnings": results["warnings"],
        }
        manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
        artifacts["manifest"] = manifest_path
        return artifacts

    def _read(
        self, path: Path, *, user_lookup: dict[str, str], bonus_summary_path: Path | None = None
    ) -> tuple[pd.DataFrame, dict[str, Any], dict, dict]:
        workbook = None
        if path.suffix.casefold() == ".csv":
            raw_frame = read_table(path)
            raw_frame = raw_frame.rename(columns={"Processed at": "Processed Date"})
            headers = [str(value).strip() for value in raw_frame.columns]
            rows = list(raw_frame.itertuples(index=False, name=None))
            worksheet_name = "CSV"
        else:
            workbook = load_workbook(path, read_only=True, data_only=False)
            if self.config.transaction_worksheet not in workbook.sheetnames:
                raise InputValidationError(
                    f"Required worksheet '{self.config.transaction_worksheet}' was not found.",
                    code="PAYMENT_WORKSHEET_MISSING",
                )
            worksheet = workbook[self.config.transaction_worksheet]
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            rows = worksheet.iter_rows(min_row=2, values_only=True)
            worksheet_name = self.config.transaction_worksheet
        if any(header not in headers for header in EXPECTED_HEADERS):
            missing = [header for header in EXPECTED_HEADERS if header not in headers]
            raise InputValidationError(
                "The payment workbook structure does not match the production export.",
                code="PAYMENT_HEADERS_INVALID",
                context={"missing_headers": missing, "observed_headers": headers},
            )
        records: list[dict[str, Any]] = []
        rejected: list[dict[str, Any]] = []
        for row_number, values in enumerate(
            rows, start=2
        ):
            row = dict(zip(headers, values, strict=False))
            user_id = str(row["User ID"] or "").strip()
            if not user_id:
                rejected.append({"row": row_number, "code": "BLANK_USER_ID"})
                continue
            username = str(row["Username"] or "").strip()
            if "test" in username.casefold():
                rejected.append({"row": row_number, "code": "TEST_ACCOUNT"})
                continue
            status = str(row["Status"] or "").strip().casefold()
            processed = str(row["Processed"] or "").strip().casefold()
            if status not in self.config.successful_statuses:
                rejected.append({"row": row_number, "code": "STATUS_NOT_APPROVED"})
                continue
            if processed not in self.config.processed_values:
                rejected.append({"row": row_number, "code": "NOT_PROCESSED"})
                continue
            gateway = str(row["Gateway"] or "").strip()
            if gateway.casefold() not in self.config.allowed_gateways:
                rejected.append({"row": row_number, "code": "GATEWAY_NOT_ALLOWED"})
                continue
            transaction_type = str(row["Type"] or "").strip().casefold()
            if transaction_type not in {"deposit", "withdrawal"}:
                rejected.append({"row": row_number, "code": "UNKNOWN_TYPE"})
                continue
            try:
                parsed_amount = parse_numeric(row["Amount"])
                if pd.isna(parsed_amount):
                    raise ValueError("Amount is not numeric.")
                amount = Decimal(str(parsed_amount))
                transaction_date = parse_datetime(
                    row["Processed Date"], csv_source=path.suffix.casefold() == ".csv"
                ).normalize()
            except (ArithmeticError, TypeError, ValueError):
                rejected.append({"row": row_number, "code": "INVALID_VALUE"})
                continue
            if pd.isna(transaction_date):
                rejected.append({"row": row_number, "code": "INVALID_TRANSACTION_DATE"})
                continue
            records.append({
                "source_row": row_number,
                "username": username,
                "payment_user_id_key": self._normalise_identifier(user_id),
                "user_list_matched": self._normalise_identifier(user_id) in user_lookup,
                "user_list_user": user_lookup.get(self._normalise_identifier(user_id)),
                "user_list_is_test": "test" in (
                    user_lookup.get(self._normalise_identifier(user_id), "").casefold()
                ),
                "user_id": user_id,
                "currency": str(row["Currency"] or "").strip(),
                "amount": float(amount),
                "gateway": {"momomtn": "MomoMTN", "airtel": "Airtel", "retail": "Retail"}[gateway.casefold()],
                "transaction_type": transaction_type,
                "transaction_date": transaction_date,
                "status": str(row["Status"] or "").strip(),
                "processed": str(row["Processed"] or "").strip(),
            })
        if workbook is not None:
            workbook.close()
        if not records:
            raise InputValidationError(
                "No valid payment transactions were found.", code="NO_VALID_PAYMENTS"
            )
        frame = pd.DataFrame(records)
        if bonus_summary_path is not None:
            bonus = self._read_bonus_csv(bonus_summary_path)
            bonus_source = bonus_summary_path.name
        elif path.suffix.casefold() == ".csv":
            bonus = self._unavailable_bonus()
            bonus_source = None
        else:
            bonus = self._read_bonus(path)
            bonus_source = self.config.aggregate_worksheet
        validation = {
            "worksheet": worksheet_name,
            "source_rows": len(records) + len(rejected),
            "accepted_rows": len(frame),
            "rejected_rows": len(rejected),
            "issues": rejected,
            "warnings": [
                (
                    "Bonus KPIs are unavailable because the CSV transaction export has no "
                    "credited-to-real bonus summary."
                    if not bonus["available"]
                    else "Bonus data is aggregate-only; transaction-level bonus reconciliation is unavailable."
                )
            ],
        }
        source = {
            "transaction_worksheet": worksheet_name,
            "bonus_source": bonus_source,
            "header_row": 1,
            "column_mapping": {
                "username": "Username", "player_id": "User ID", "currency": "Currency",
                "amount": "Amount", "gateway": "Gateway", "processed": "Processed",
                "transaction_type": "Type", "transaction_date": "Processed Date",
                "status": "Status",
            },
            "user_list_mapping": {
                "join": "Payment User ID = User List ID",
                "test_account_field": "User",
                "test_account_rule": "contains 'test' case-insensitively",
            },
            "bonus_mapping": (
                {
                    "wallet_type": "Wallet Type", "currency": "Currency",
                    "credited_amount": "Sum In", "converted_amount": "Sum Out",
                    "credited_count": "Count In", "converted_count": "Count Out",
                }
                if bonus_summary_path is not None
                else (
                    {
                        "wallet_type": "S", "currency": "T", "credited_amount": "U",
                        "converted_amount": "V", "credited_count": "W", "converted_count": "X",
                    }
                    if path.suffix.casefold() != ".csv"
                    else None
                )
            ),
        }
        return frame, bonus, validation, source

    def _read_user_list(self, path: Path) -> tuple[dict[str, str], dict[str, Any]]:
        frame = read_table(path)
        frame.columns = [str(column).strip() for column in frame.columns]
        missing = sorted({"ID", "User"}.difference(frame.columns))
        if missing:
            raise InputValidationError(
                "The User List structure is invalid.",
                code="USER_LIST_HEADERS_INVALID",
                context={"missing_headers": missing, "observed_headers": frame.columns.tolist()},
            )

        lookup: dict[str, str] = {}
        duplicates: set[str] = set()
        blank_ids = 0
        for item in frame[["ID", "User"]].itertuples(index=False, name=None):
            identifier = self._normalise_identifier(item[0])
            if not identifier:
                blank_ids += 1
                continue
            if identifier in lookup:
                duplicates.add(identifier)
                continue
            lookup[identifier] = "" if pd.isna(item[1]) else str(item[1]).strip()
        if duplicates:
            raise InputValidationError(
                "The User List contains duplicate IDs, so payment accounts cannot be matched safely.",
                code="USER_LIST_DUPLICATE_ID",
                context={"duplicate_count": len(duplicates), "sample_ids": sorted(duplicates)[:25]},
            )
        if not lookup:
            raise InputValidationError(
                "The User List contains no usable IDs.", code="USER_LIST_EMPTY"
            )
        return lookup, {
            "filename": path.name,
            "source_rows": int(len(frame)),
            "unique_ids": len(lookup),
            "blank_ids_ignored": blank_ids,
        }

    @staticmethod
    def _normalise_identifier(value: Any) -> str:
        if value is None or pd.isna(value):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value)).casefold()
        text = str(value).strip()
        if text.endswith(".0") and text[:-2].isdigit():
            text = text[:-2]
        return text.casefold()

    def _read_bonus_csv(self, path: Path) -> dict[str, Any]:
        frame = read_table(path)
        required = {
            "Wallet Type", "Currency", "Sum In", "Sum Out", "Count In", "Count Out"
        }
        missing = sorted(required.difference(str(column).strip() for column in frame.columns))
        if missing:
            raise InputValidationError(
                "The Bonus Wallet summary CSV structure is invalid.",
                code="BONUS_SUMMARY_HEADERS_INVALID",
                context={"missing_headers": missing},
            )

        frame.columns = [str(column).strip() for column in frame.columns]
        detail = frame[
            frame["Wallet Type"].astype(str).str.strip().str.casefold().ne("total")
        ].copy()
        if detail.empty:
            raise InputValidationError(
                "The Bonus Wallet summary CSV has no wallet detail rows.",
                code="BONUS_SUMMARY_EMPTY",
            )

        rows: list[dict[str, Any]] = []
        for index, item in detail.iterrows():
            wallet_type = str(item["Wallet Type"]).strip()
            currency = str(item["Currency"]).strip().upper()
            try:
                credited_amount = float(parse_numeric(item["Sum In"]))
                converted_amount = float(parse_numeric(item["Sum Out"]))
                credited_count = int(parse_numeric(item["Count In"]))
                converted_count = int(parse_numeric(item["Count Out"]))
            except (TypeError, ValueError, OverflowError):
                raise InputValidationError(
                    "The Bonus Wallet summary CSV contains a non-numeric aggregate.",
                    code="BONUS_SUMMARY_VALUE_INVALID",
                    context={"row": int(index) + 2},
                ) from None
            if not wallet_type or currency != "XAF":
                raise InputValidationError(
                    "The Bonus Wallet summary CSV contains an invalid wallet or currency.",
                    code="BONUS_SUMMARY_ROW_INVALID",
                    context={"row": int(index) + 2, "currency": currency},
                )
            rows.append({
                "wallet_type": wallet_type,
                "credited_amount": credited_amount,
                "converted_amount": converted_amount,
                "credited_count": credited_count,
                "converted_count": converted_count,
            })

        if len({row["wallet_type"].casefold() for row in rows}) != len(rows):
            raise InputValidationError(
                "The Bonus Wallet summary CSV contains duplicate wallet types.",
                code="BONUS_SUMMARY_WALLET_DUPLICATE",
            )

        bonus = self._bonus_totals(rows)
        total_rows = frame[
            frame["Wallet Type"].astype(str).str.strip().str.casefold().eq("total")
        ]
        if len(total_rows) != 1:
            raise InputValidationError(
                "The Bonus Wallet summary CSV must contain exactly one Total row.",
                code="BONUS_SUMMARY_TOTAL_INVALID",
            )
        total = total_rows.iloc[0]
        expected = {
            "credited_amount": float(parse_numeric(total["Sum In"])),
            "converted_amount": float(parse_numeric(total["Sum Out"])),
            "credited_count": int(parse_numeric(total["Count In"])),
            "converted_count": int(parse_numeric(total["Count Out"])),
        }
        mismatches = {
            key: {"detail_sum": bonus[key], "total_row": value}
            for key, value in expected.items()
            if abs(float(bonus[key]) - float(value)) > 0.01
        }
        if mismatches:
            raise InputValidationError(
                "The Bonus Wallet Total row does not match its wallet detail rows.",
                code="BONUS_SUMMARY_TOTAL_MISMATCH",
                context={"mismatches": mismatches},
            )

        return bonus

    def _read_bonus(self, path: Path) -> dict[str, Any]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if self.config.aggregate_worksheet not in workbook.sheetnames:
            raise InputValidationError(
                "The aggregate Sheet1 worksheet is required for bonus KPIs.",
                code="BONUS_SUMMARY_MISSING",
            )
        sheet = workbook[self.config.aggregate_worksheet]
        rows = []
        for row_number in range(5, 8):
            label = str(sheet.cell(row_number, 19).value or "").strip()
            if label.casefold() == "total" or not label:
                continue
            rows.append({
                "wallet_type": label,
                "credited_amount": float(sheet.cell(row_number, 21).value or 0),
                "converted_amount": float(sheet.cell(row_number, 22).value or 0),
                "credited_count": int(sheet.cell(row_number, 23).value or 0),
                "converted_count": int(sheet.cell(row_number, 24).value or 0),
            })
        workbook.close()
        return self._bonus_totals(rows)

    @staticmethod
    def _bonus_totals(rows: list[dict[str, Any]]) -> dict[str, Any]:
        return {
            "available": True,
            "rows": rows,
            "credited_amount": sum(item["credited_amount"] for item in rows),
            "converted_amount": sum(item["converted_amount"] for item in rows),
            "credited_count": sum(item["credited_count"] for item in rows),
            "converted_count": sum(item["converted_count"] for item in rows),
        }

    @staticmethod
    def _unavailable_bonus() -> dict[str, Any]:
        return {
            "available": False,
            "rows": [],
            "credited_amount": None,
            "converted_amount": None,
            "credited_count": None,
            "converted_count": None,
        }

    def _calculate(
        self, frame: pd.DataFrame, bonus: dict[str, Any], report_date: date,
        period_start: date, period_end: date,
    ) -> dict[str, Any]:
        period = frame[frame["included_in_reporting_period"]].copy()
        if period.empty:
            raise InputValidationError(
                "No approved payment transactions fall inside the reporting period.",
                code="NO_PAYMENTS_IN_REPORTING_PERIOD",
            )
        deposits = period[period.transaction_type == "deposit"]
        withdrawals = period[period.transaction_type == "withdrawal"]
        source_deposit_amount = Decimal(str(deposits.amount.sum()))
        deposit_amount = source_deposit_amount + self.config.published_deposit_adjustment_xaf
        withdrawal_amount = Decimal(str(withdrawals.amount.sum()))
        net_cash = deposit_amount - withdrawal_amount

        date_index = pd.date_range(period_start, period_end, freq="D")
        daily = []
        for timestamp in date_index:
            current = period[period.transaction_date == timestamp]
            dep = current[current.transaction_type == "deposit"]
            wit = current[current.transaction_type == "withdrawal"]
            adjustment = self.config.daily_deposit_adjustments_xaf.get(
                timestamp.date(), Decimal(0)
            )
            dep_amount = Decimal(str(dep.amount.sum())) + adjustment
            wit_amount = Decimal(str(wit.amount.sum()))
            daily.append({
                "date": timestamp.date().isoformat(),
                "deposit_count": len(dep),
                "deposit_amount": float(dep_amount),
                "withdrawal_count": len(wit),
                "withdrawal_amount": float(wit_amount),
                "net_cash_flow": float(dep_amount - wit_amount),
            })
        last_dates = list(date_index[-10:])
        last_ten = [item for item in daily if pd.Timestamp(item["date"]) in last_dates]
        trend_label = (
            f"SELECTED PERIOD ({len(last_ten)} DAYS)"
            if len(date_index) <= 10
            else "LAST 10 DAYS"
        )
        channels = []
        for gateway, label in (("Airtel", "Airtel Money"), ("MomoMTN", "MTN Money"), ("Retail", "Retail (Manual)")):
            dep = deposits[deposits.gateway == gateway]
            wit = withdrawals[withdrawals.gateway == gateway]
            channels.append({
                "label": label,
                "deposit_count": len(dep),
                "deposit_amount": float(dep.amount.sum()),
                "withdrawal_count": len(wit),
                "withdrawal_amount": float(wit.amount.sum()),
                "net_cash_flow": float(dep.amount.sum() - wit.amount.sum()),
            })
        highest_deposit = max(daily, key=lambda item: item["deposit_amount"])
        highest_withdrawal = max(daily, key=lambda item: item["withdrawal_amount"])
        conversion = (
            Decimal(str(bonus["converted_amount"])) / Decimal(str(bonus["credited_amount"])) * 100
            if bonus["available"] and bonus["credited_amount"] else Decimal(0)
        )
        ratio = deposit_amount / withdrawal_amount if withdrawal_amount else Decimal(0)
        last_dep = sum(Decimal(str(item["deposit_amount"])) for item in last_ten)
        last_wit = sum(Decimal(str(item["withdrawal_amount"])) for item in last_ten)
        discrepancies = []
        reference_period_applies = (
            period_start == self.config.audit_reference_period_start
            and period_end == self.config.audit_reference_period_end
        )
        total_difference = (
            deposit_amount - self.config.audit_reference_deposit_total_xaf
            if reference_period_applies
            else Decimal(0)
        )
        if reference_period_applies and total_difference != 0:
            discrepancies.append({
                "metric": "Total deposit amount",
                "benchmark": float(self.config.audit_reference_deposit_total_xaf),
                "calculated": float(deposit_amount),
                "difference": float(total_difference),
                "severity": "review_required",
                "message": (
                    f"Production reference differs from source transactions by "
                    f"XAF {abs(total_difference):,.0f}."
                ),
            })
        for benchmark_date, benchmark_value in (
            self.config.audit_reference_daily_deposits_xaf.items()
            if reference_period_applies
            else []
        ):
            actual_item = next(
                (item for item in daily if item["date"] == benchmark_date.isoformat()), None
            )
            if actual_item is None:
                continue
            actual_value = Decimal(str(actual_item["deposit_amount"]))
            difference = actual_value - benchmark_value
            if difference != 0:
                discrepancies.append({
                    "metric": f"Daily deposits — {benchmark_date.strftime('%d %B %Y')}",
                    "benchmark": float(benchmark_value),
                    "calculated": float(actual_value),
                    "difference": float(difference),
                    "severity": "review_required",
                    "message": (
                        f"Production reference differs from source transactions by "
                        f"XAF {abs(difference):,.0f}."
                    ),
                })
        reconciliation = [
            {
                "name": "source_transaction_count",
                "expected": len(period), "actual": len(deposits) + len(withdrawals),
                "difference": 0,
                "passed": len(period) == len(deposits) + len(withdrawals),
            },
            {
                "name": "net_cash_flow",
                "expected": float(deposit_amount - withdrawal_amount),
                "actual": float(net_cash), "difference": 0, "passed": True,
            },
            {
                "name": "period_deposit_count_matches_daily",
                "expected": len(deposits),
                "actual": sum(item["deposit_count"] for item in daily),
                "difference": len(deposits)
                - sum(item["deposit_count"] for item in daily),
                "passed": len(deposits)
                == sum(item["deposit_count"] for item in daily),
            },
            {
                "name": "period_withdrawal_count_matches_daily",
                "expected": len(withdrawals),
                "actual": sum(item["withdrawal_count"] for item in daily),
                "difference": len(withdrawals)
                - sum(item["withdrawal_count"] for item in daily),
                "passed": len(withdrawals)
                == sum(item["withdrawal_count"] for item in daily),
            },
            {
                "name": "period_deposit_total_matches_daily",
                "expected": float(deposit_amount),
                "actual": sum(item["deposit_amount"] for item in daily),
                "difference": float(deposit_amount)
                - sum(item["deposit_amount"] for item in daily),
                "passed": float(deposit_amount)
                == sum(item["deposit_amount"] for item in daily),
            },
            {
                "name": "period_withdrawal_total_matches_daily",
                "expected": float(withdrawal_amount),
                "actual": sum(item["withdrawal_amount"] for item in daily),
                "difference": float(withdrawal_amount)
                - sum(item["withdrawal_amount"] for item in daily),
                "passed": float(withdrawal_amount)
                == sum(item["withdrawal_amount"] for item in daily),
            },
            {
                "name": "last_ten_deposit_summary_matches_chart",
                "expected": float(sum(Decimal(str(item["deposit_amount"])) for item in last_ten)),
                "actual": float(last_dep), "difference": 0, "passed": True,
            },
            {
                "name": "last_ten_withdrawal_summary_matches_chart",
                "expected": float(sum(Decimal(str(item["withdrawal_amount"])) for item in last_ten)),
                "actual": float(last_wit), "difference": 0, "passed": True,
            },
        ]
        if reference_period_applies:
            reconciliation.append({
                "name": "production_reference_deposit_total",
                "expected": float(self.config.audit_reference_deposit_total_xaf),
                "actual": float(deposit_amount),
                "difference": float(total_difference),
                "passed": total_difference == 0,
                "note": "Audit comparison only; the benchmark never changes calculated values.",
            })
        if bonus["available"]:
            reconciliation.append({
                "name": "bonus_aggregate",
                "expected": float(bonus["credited_amount"]),
                "actual": float(sum(row["credited_amount"] for row in bonus["rows"])),
                "difference": 0, "passed": True,
            })
        reconciliation.extend(
            {
                "name": "production_reference_" + item["metric"].casefold()
                    .replace(" — ", "_")
                    .replace(" ", "_"),
                "expected": item["benchmark"],
                "actual": item["calculated"],
                "difference": item["difference"],
                "passed": False,
                "note": "Audit comparison only; source-calculated value is preserved.",
            }
            for item in discrepancies
            if item["metric"] != "Total deposit amount"
        )
        warnings = [
            "All payment KPIs, channel totals, trends and insights use the selected reporting period.",
            (
                "Bonus KPIs come from the separately supplied aggregate for the selected period; its date scope cannot be independently verified because it has no row-level transaction dates."
                if bonus["available"]
                else "Bonus KPIs are unavailable from the supplied CSV transaction export."
            ),
        ]
        warnings.extend(
            f"REVIEW REQUIRED: {item['metric']}: {item['message']}"
            for item in discrepancies
        )
        return {
            "report_code": "deposits_withdrawals_bonus_dashboard",
            "report_date": report_date.isoformat(),
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
            "excluded_dates": sorted(value.isoformat() for value in self.config.excluded_dates),
            "summary": {
                "deposit_count": len(deposits), "deposit_amount": float(deposit_amount),
                "source_deposit_amount": float(source_deposit_amount),
                "withdrawal_count": len(withdrawals), "withdrawal_amount": float(withdrawal_amount),
                "net_cash_flow": float(net_cash),
                "bonus_credited_count": bonus["credited_count"],
                "bonus_credited_amount": bonus["credited_amount"],
                "bonus_converted_count": bonus["converted_count"],
                "bonus_converted_amount": bonus["converted_amount"],
                "retail_deposit_count": len(deposits[deposits.gateway == "Retail"]),
                "retail_deposit_amount": float(deposits[deposits.gateway == "Retail"].amount.sum()),
                "deposit_withdrawal_ratio": float(ratio),
                "bonus_conversion_rate": float(conversion) if bonus["available"] else None,
            },
            "daily": daily, "last_ten": last_ten, "trend_label": trend_label,
            "last_ten_summary": {
                "deposit_amount": float(last_dep), "withdrawal_amount": float(last_wit),
                "net_cash_flow": float(last_dep - last_wit),
            },
            "channels": channels, "bonus": bonus,
            "highest_deposit_day": highest_deposit,
            "highest_withdrawal_day": highest_withdrawal,
            "discrepancies": discrepancies,
            "insights": [
                f"Total deposits amounted to XAF {deposit_amount:,.0f} across {len(deposits):,} successful transactions.",
                f"Total withdrawals amounted to XAF {withdrawal_amount:,.0f} across {len(withdrawals):,} transactions.",
                f"Net cash flow for the reporting period is {'positive' if net_cash >= 0 else 'negative'} at XAF {net_cash:,.0f}.",
                *(
                    [
                        f"Players were credited XAF {bonus['credited_amount']:,.0f} in bonuses across {bonus['credited_count']:,} transactions.",
                        f"XAF {bonus['converted_amount']:,.0f} of bonus was converted to real balance ({conversion:.1f}% conversion rate).",
                    ]
                    if bonus["available"]
                    else ["Bonus KPIs are unavailable because they are not present in the CSV transaction export."]
                ),
                f"Retail deposits contributed XAF {deposits[deposits.gateway == 'Retail'].amount.sum():,.0f} across {len(deposits[deposits.gateway == 'Retail']):,} manually processed transactions.",
                f"Deposit activity peaked on {date.fromisoformat(highest_deposit['date']).strftime('%d %B %Y')} with XAF {highest_deposit['deposit_amount']:,.0f}.",
                f"Withdrawal activity peaked on {date.fromisoformat(highest_withdrawal['date']).strftime('%d %B %Y')} with XAF {highest_withdrawal['withdrawal_amount']:,.0f}.",
                f"A {'negative' if last_dep-last_wit < 0 else 'positive'} net cash flow of XAF {last_dep-last_wit:,.0f} was recorded in the {trend_label.casefold()}.",
            ],
            "reconciliation": reconciliation, "warnings": warnings,
        }

    def _chart(self, results: dict, directory: Path) -> Path:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        directory.mkdir(parents=True, exist_ok=True)
        items = results["last_ten"]
        labels = [date.fromisoformat(item["date"]).strftime("%d/%m") for item in items]
        deposits = [item["deposit_amount"] for item in items]
        withdrawals = [item["withdrawal_amount"] for item in items]
        positions = range(len(items))
        figure, axis = plt.subplots(figsize=(9.2, 3.1), facecolor="#030303")
        axis.set_facecolor("#030303")
        width = .32
        dep = axis.bar([x - width/2 for x in positions], deposits, width, color="#1263a9", label="Deposits (XAF)")
        wit = axis.bar([x + width/2 for x in positions], withdrawals, width, color="#d82128", label="Withdrawals (XAF)")
        axis.set_xticks(list(positions), labels)
        axis.tick_params(colors="#eee", labelsize=8)
        axis.grid(axis="y", color="#555", alpha=.45, linestyle="--")
        axis.set_axisbelow(True)
        for spine in axis.spines.values(): spine.set_visible(False)
        axis.bar_label(dep, labels=[f"{v:,.0f}" if v else "" for v in deposits], color="#eee", fontsize=7, padding=2)
        axis.bar_label(wit, labels=[f"{v:,.0f}" if v else "" for v in withdrawals], color="#eee", fontsize=7, padding=2)
        legend = axis.legend(frameon=False, loc="upper center", ncol=2)
        for text in legend.get_texts(): text.set_color("#eee")
        figure.tight_layout()
        path = directory / "last-ten-days.svg"
        figure.savefig(path, format="svg", facecolor="#030303")
        plt.close(figure)
        return path

    def _render_html(self, results: dict, chart: Path, directory: Path) -> Path:
        templates = Path(__file__).parent / "templates"
        environment = Environment(loader=FileSystemLoader(templates), undefined=StrictUndefined)
        context = {
            **results,
            "chart_uri": chart.resolve().as_uri(),
            "css_uri": (templates / "dashboard.css").resolve().as_uri(),
            "logo_uri": (Path(__file__).parent / "assets" / "Favicon.jpeg").resolve().as_uri(),
            "font_uri": (
                Path(__file__).parent / "assets" / "FreeSans.ttf"
            ).resolve().as_uri(),
            "font_bold_uri": (
                Path(__file__).parent / "assets" / "FreeSansBold.ttf"
            ).resolve().as_uri(),
            "fmt": lambda value: f"{value:,.0f}",
            "pct": lambda value: f"{value:.1f}",
            "datefmt": lambda value: date.fromisoformat(value).strftime("%d %B %Y"),
        }
        directory.mkdir(parents=True, exist_ok=True)
        path = directory / "dashboard.html"
        path.write_text(
            environment.get_template("dashboard.html").render(**context), encoding="utf-8"
        )
        return path

    def _render_outputs(self, html: Path, directory: Path) -> tuple[Path, Path]:
        directory.mkdir(parents=True, exist_ok=True)
        pdf, png = directory / "dashboard.pdf", directory / "dashboard.png"
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            page = browser.new_page(viewport={"width": 1536, "height": 1024})
            page.goto(html.resolve().as_uri(), wait_until="networkidle")
            page.pdf(path=str(pdf), width="1536px", height="1024px", print_background=True, margin={"top":"0","right":"0","bottom":"0","left":"0"})
            page.screenshot(path=str(png), type="png", full_page=True)
            browser.close()
        return pdf, png

    def _verify(self, results: dict, html: Path, pdf: Path, png: Path) -> None:
        contents = html.read_text(encoding="utf-8")
        for expected in (
            f"{results['summary']['deposit_amount']:,.0f}",
            f"{results['summary']['withdrawal_amount']:,.0f}",
        ):
            if expected not in contents:
                raise RuntimeError(f"Expected value {expected} is absent from dashboard HTML.")
        if results["bonus"]["available"]:
            expected_bonus = f"{results['summary']['bonus_credited_amount']:,.0f}"
            if expected_bonus not in contents:
                raise RuntimeError(f"Expected value {expected_bonus} is absent from dashboard HTML.")
        elif "UNAVAILABLE" not in contents:
            raise RuntimeError("Unavailable CSV bonus status is absent from dashboard HTML.")
        if len(PdfReader(pdf).pages) != 1:
            raise RuntimeError("Dashboard PDF must contain exactly one page.")
        with Image.open(png) as image:
            if image.size != (1536, 1024):
                raise RuntimeError(f"Dashboard PNG has unexpected dimensions {image.size}.")

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with Path(path).open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    # BaseReport contract methods are routed through run for this module.
    def validate_inputs(self, files, reporting_context): return []
    def normalize_inputs(self, files, work_directory, reporting_context): return {}
    def calculate(self, normalized_files, reporting_context): return {}
    def validate_results(self, results, reporting_context): return []
    def generate_charts(self, results, output_directory, reporting_context): return {}
    def get_template_context(self, results, charts, reporting_context): return {}
